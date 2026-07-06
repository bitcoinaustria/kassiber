"""File-format parsers for wallet-export imports.

Call sites hand a file path + format tag in and get back a list of
normalized record dicts out — one dict per transaction, with the same
shape across all formats so downstream DB-insertion code
(`insert_wallet_records` in `kassiber.core.imports`) doesn't need to
branch per wallet.

Import format families live here:

  - Generic JSON / CSV — user-supplied, pass-through (no asset-specific
    parsing beyond CSV-row decoding).
  - BTCPay Server (`btcpay_json`, `btcpay_csv`) — on-chain receive/send
    history exported from a BTCPay store. `normalize_btcpay_record`
    pulls the `_btcpay_comment` and `_btcpay_labels` fields so
    `apply_btcpay_metadata` can later set transaction notes and tags.
  - Phoenix (`phoenix_csv`) — Lightning-native CSV from the Phoenix
    mobile wallet. Amounts are stored as signed INTEGER msat, fees are
    split across `mining_fee_sat` + `service_fee_msat`, and
    `amount_fiat` is "value CCY" formatted. We normalize all of it to
    the common record shape (BTC Decimals + derived `fiat_rate`).
  - River (`river_csv`) — Bitcoin Activity or Account Activity CSV
    exports. BTC-side rows are normalized to Kassiber transactions and
    exact fiat execution amounts are preserved as exchange pricing
    provenance where the export provides the paired cash leg.
  - Bull Bitcoin (`bullbitcoin_csv`) — order export CSV. Completed
    Bitcoin on-chain, Lightning, or Liquid <-> fiat orders are normalized
    to exchange execution pricing, keyed by the exported transaction id
    when present.
  - Bull Bitcoin wallet (`bullbitcoin_wallet_csv`) — unified wallet-history
    CSV. Bitcoin, Liquid, Lightning, payjoin, and swap rows are normalized as
    wallet activity without fiat execution pricing.
  - Coinfinity (`coinfinity_csv`) — order export CSV. BTC/EUR broker rows
    are normalized as exact exchange execution evidence, with fiat fees
    folded into the taxable buy/sell value.
  - 21bitcoin (`21bitcoin_csv`) — transaction CSV export. BTC-side trades
    and withdrawals are normalized as a custodial platform ledger; fiat-only
    cash rows are skipped.
  - Pocket Bitcoin (`pocketbitcoin_csv`) — account CSV export. Exchange rows
    are normalized as exact execution evidence and paired with adjacent BTC
    withdrawal rows when the export records the on-chain payout separately.
  - Strike (`strike_csv`) — custodial platform CSV export. Bitcoin Lightning,
    on-chain, and BTC exchange rows are normalized as active platform ledger
    rows; fiat-only funding rows are skipped.
  - BIP329 JSONL — one-record-per-line label export. `record_type`
    distinguishes tx/addr/pubkey/input/output/xpub labels.

`load_import_records(path, input_format)` is the dispatcher used by the
coordinator layer. Adding a new format: add a `normalize_<fmt>_record`
and a `load_<fmt>_records`, plug them into `load_import_records`, and
add an `is_<fmt>_format` predicate so the metadata-application layer in
`kassiber.core.imports` can hook format-specific side effects.

Every parser raises `AppError` on unparseable input so the CLI surfaces
a validation envelope rather than a bare `ValueError` / `KeyError`.
"""

import csv
import hashlib
import io
import json
import os
import re
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Any, Mapping

from .envelope import json_ready
from .errors import AppError
from .msat import dec, msat_to_btc
from .time_utils import parse_timestamp
from .util import str_or_none
from .wallet_descriptors import normalize_asset_code


# -- generic coordinator -----------------------------------------------------


def load_import_records(file_path, input_format):
    """Dispatch on `input_format` to the matching loader.

    Returns a list of dicts with a format-appropriate shape. Generic
    JSON / CSV pass through raw rows; BTCPay and Phoenix loaders
    return already-normalized records ready for
    `normalize_import_record`.
    """
    if not os.path.exists(file_path):
        raise AppError(
            f"Import file not found: {file_path}",
            code="not_found",
            hint="Check the wallet source_file or pass an existing --file path.",
        )
    if input_format == "json":
        with open(file_path, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
        if isinstance(payload, dict):
            payload = payload.get("transactions", [])
        if not isinstance(payload, list):
            raise AppError("JSON import must be a list of transaction objects")
        return payload
    if input_format == "csv":
        with open(file_path, "r", encoding="utf-8", newline="") as handle:
            return list(csv.DictReader(handle))
    if input_format == "btcpay_json":
        return load_btcpay_export_records(file_path, "json")
    if input_format == "btcpay_csv":
        return load_btcpay_export_records(file_path, "csv")
    if input_format == "phoenix_csv":
        return load_phoenix_csv_records(file_path)
    if input_format == "river_csv":
        return load_river_csv_records(file_path)
    if input_format == "bullbitcoin_csv":
        return load_bullbitcoin_csv_records(file_path)
    if input_format == "bullbitcoin_wallet_csv":
        return load_bullbitcoin_wallet_csv_records(file_path)
    if input_format == "coinfinity_csv":
        return load_coinfinity_csv_records(file_path)
    if input_format == "21bitcoin_csv":
        return load_twentyonebitcoin_csv_records(file_path)
    if input_format == "pocketbitcoin_csv":
        return load_pocketbitcoin_csv_records(file_path)
    if input_format == "strike_csv":
        return load_strike_csv_records(file_path)
    if input_format == "ledgerlive_csv":
        return load_ledgerlive_csv_records(file_path)
    if input_format == "binance_supplemental_csv":
        return load_binance_supplemental_csv_records(file_path)
    if input_format == "wasabi_bundle":
        return load_wasabi_bundle_records(file_path)
    if input_format == GENERIC_LEDGER_FORMAT:
        return load_generic_ledger_records(file_path)
    raise AppError(f"Unsupported input format '{input_format}'")


# -- Wasabi Wallet -----------------------------------------------------------


WASABI_BUNDLE_FORMAT = "wasabi_bundle"
_WASABI_SENSITIVE_KEY_NAMES = {
    "accountkeypath",
    "chaincode",
    "encryptedsecret",
    "encryptedsecretdecryptioninfos",
    "extpubkey",
    "fullkeypath",
    "keypath",
    "masterkeyfingerprint",
    "password",
    "passphrase",
    "privatekey",
    "pubkey",
    "publickey",
    "walletfile",
    "xpub",
}


def _wasabi_bundle_key(value: Any) -> str:
    return str(value or "").replace("-", "").replace("_", "").replace(".", "").casefold()


def _wasabi_result(value: Any) -> Any:
    if isinstance(value, dict) and "result" in value:
        return value.get("result")
    return value


def _wasabi_section(payload: Any, *names: str) -> Any:
    if not isinstance(payload, dict):
        return None
    normalized_names = {_wasabi_bundle_key(name) for name in names}
    for key, value in payload.items():
        if _wasabi_bundle_key(key) in normalized_names:
            return _wasabi_result(value)
    for container_name in ("rpc", "responses", "exports", "wasabi"):
        nested = payload.get(container_name)
        if isinstance(nested, dict):
            value = _wasabi_section(nested, *names)
            if value is not None:
                return value
    return None


def _wasabi_section_present(payload: Any, *names: str) -> bool:
    if not isinstance(payload, dict):
        return False
    normalized_names = {_wasabi_bundle_key(name) for name in names}
    for key in payload:
        if _wasabi_bundle_key(key) in normalized_names:
            return True
    for container_name in ("rpc", "responses", "exports", "wasabi"):
        nested = payload.get(container_name)
        if isinstance(nested, dict) and _wasabi_section_present(nested, *names):
            return True
    return False


def _wasabi_list(payload: Any, *names: str) -> list[dict[str, Any]]:
    value = _wasabi_section(payload, *names)
    if value is None:
        return []
    if isinstance(value, dict):
        for key in ("items", "records", "transactions", "history", "coins", "keys", "payments"):
            nested = value.get(key)
            if isinstance(nested, list):
                value = nested
                break
    if not isinstance(value, list):
        raise AppError(f"Wasabi bundle section '{names[0]}' must be a list", code="validation")
    return [item for item in value if isinstance(item, dict)]


def _wasabi_dict(payload: Any, *names: str) -> dict[str, Any]:
    value = _wasabi_section(payload, *names)
    return value if isinstance(value, dict) else {}


def _wasabi_pick(mapping: Mapping[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in mapping and mapping[key] not in (None, ""):
            return mapping[key]
    return None


def _wasabi_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if value is None or value == "":
        return default
    text = str(value).strip().casefold()
    if text in {"1", "true", "yes", "on"}:
        return True
    if text in {"0", "false", "no", "off"}:
        return False
    return default


def _wasabi_int(value: Any, field: str, *, default: int | None = None) -> int | None:
    if value is None or value == "":
        return default
    try:
        if isinstance(value, Decimal):
            return int(value)
        if isinstance(value, float):
            return int(value)
        text = str(value).strip()
        if "." in text:
            return int(Decimal(text))
        return int(text)
    except (TypeError, ValueError, ArithmeticError) as exc:
        raise AppError(f"Wasabi bundle has invalid {field}", code="validation") from exc


def _wasabi_amount_sats(value: Any) -> int:
    if value is None or value == "":
        raise AppError("Wasabi transaction is missing amount", code="validation")
    if isinstance(value, int):
        return value
    text = str(value).strip()
    try:
        if "." in text:
            return int((dec(text) * Decimal("100000000")).to_integral_value())
        return int(text)
    except (ValueError, ArithmeticError) as exc:
        raise AppError(f"Wasabi bundle has invalid amount '{value}'", code="validation") from exc


def _wasabi_btc_from_sats(value: int) -> Decimal:
    return Decimal(value) / Decimal("100000000")


def _wasabi_safe_txid(value: Any) -> str | None:
    text = str_or_none(value)
    if text is None:
        return None
    text = text.strip().lower()
    if len(text) != 64:
        return text
    try:
        bytes.fromhex(text)
    except ValueError:
        return text
    return text


def _wasabi_path_parts(value: Any) -> list[str]:
    text = str_or_none(value)
    if text is None:
        return []
    parts = [part.strip() for part in text.split("/") if part.strip()]
    if parts and parts[0].lower() == "m":
        parts = parts[1:]
    return parts


def _wasabi_key_path_tail(value: Any) -> dict[str, Any]:
    parts = _wasabi_path_parts(value)
    if not parts:
        return {"branch_label": None, "branch_index": None, "address_index": None, "has_key_path": False}
    if len(parts) < 2:
        return {"branch_label": None, "branch_index": None, "address_index": None, "has_key_path": True}
    try:
        branch = int(parts[-2].rstrip("'hH"))
        index = int(parts[-1].rstrip("'hH"))
    except ValueError:
        return {"branch_label": None, "branch_index": None, "address_index": None, "has_key_path": True}
    branch_label = "receive" if branch == 0 else "change" if branch == 1 else f"branch-{branch}"
    return {
        "branch_label": branch_label,
        "branch_index": branch,
        "address_index": index,
        "has_key_path": True,
    }


def _wasabi_account_hint(value: Any) -> dict[str, Any] | None:
    if not isinstance(value, dict):
        return None
    name = str_or_none(value.get("name") or value.get("Name"))
    path = str_or_none(value.get("keyPath") or value.get("KeyPath") or value.get("fullKeyPath"))
    hint: dict[str, Any] = {"name": name or "account"}
    if path:
        parts = _wasabi_path_parts(path)
        purpose = parts[0].rstrip("'hH") if len(parts) >= 1 else ""
        coin_type = parts[1].rstrip("'hH") if len(parts) >= 2 else ""
        if purpose.isdigit():
            hint["purpose"] = int(purpose)
        if coin_type.isdigit():
            hint["coin_type"] = int(coin_type)
        if len(parts) >= 3:
            hint["account_path_hint"] = f"{parts[0]}/{parts[1]}/*'"
    return hint


def _wasabi_account_path_hint(value: Any) -> str | None:
    parts = _wasabi_path_parts(value)
    if not parts:
        return None
    if len(parts) < 2:
        return "account_path_present"
    return f"{parts[0]}/{parts[1]}/*'"


def _wasabi_sanitize_mapping(value: Any) -> Any:
    if isinstance(value, dict):
        sanitized = {}
        for key, item in value.items():
            safe_key = str(key)
            if _wasabi_bundle_key(safe_key) in _WASABI_SENSITIVE_KEY_NAMES:
                continue
            sanitized[safe_key] = _wasabi_sanitize_mapping(item)
        return sanitized
    if isinstance(value, list):
        return [_wasabi_sanitize_mapping(item) for item in value]
    return value


def _wasabi_wallet_metadata(wallet_info: dict[str, Any], wallet_json: dict[str, Any]) -> dict[str, Any]:
    merged = {**wallet_json, **wallet_info}
    if not merged:
        return {}

    def pick(*keys: str) -> Any:
        for key in keys:
            if key in merged:
                return merged[key]
        folded = {_wasabi_bundle_key(key): value for key, value in merged.items()}
        for key in keys:
            lookup = _wasabi_bundle_key(key)
            if lookup in folded:
                return folded[lookup]
        return None

    accounts_raw = pick("accounts", "Accounts") or []
    accounts = []
    if isinstance(accounts_raw, list):
        accounts = [hint for item in accounts_raw if (hint := _wasabi_account_hint(item))]
    metadata = {
        "walletName": str_or_none(pick("walletName", "Name")),
        "state": str_or_none(pick("State", "state")),
        "anonScoreTarget": _wasabi_int(pick("anonScoreTarget", "AnonScoreTarget"), "anonScoreTarget"),
        "autoCoinJoin": _wasabi_bool(pick("isAutoCoinjoin", "AutoCoinJoin"), False),
        "redCoinIsolation": _wasabi_bool(pick("RedCoinIsolation", "isNonPrivateCoinIsolation"), False),
        "isWatchOnly": _wasabi_bool(pick("isWatchOnly", "IsWatchOnly"), False),
        "isHardwareWallet": _wasabi_bool(pick("isHardwareWallet", "IsHardwareWallet"), False),
        "minGapLimit": _wasabi_int(pick("MinGapLimit", "minGapLimit"), "MinGapLimit"),
        "coinjoinStatus": str_or_none(pick("coinjoinStatus", "CoinjoinStatus")),
        "segwitAccountPathHint": _wasabi_account_path_hint(
            pick("SegWitAccountKeyPath", "segwitAccountKeyPath")
        ),
        "taprootAccountPathHint": _wasabi_account_path_hint(
            pick("TaprootAccountKeyPath", "taprootAccountKeyPath")
        ),
        "silentPaymentAccountPathHint": _wasabi_account_path_hint(
            pick("SilentPaymentAccountKeyPath", "silentPaymentAccountKeyPath")
        ),
        "accounts": accounts,
        "walletJsonImported": bool(wallet_json),
    }
    return {key: value for key, value in metadata.items() if value not in (None, "", [])}


def _normalize_wasabi_history_record(record: dict[str, Any]) -> dict[str, Any]:
    sanitized = {str(key): value for key, value in record.items() if key is not None}
    txid = _wasabi_safe_txid(
        sanitized.get("tx")
        or sanitized.get("txid")
        or sanitized.get("transactionId")
        or sanitized.get("transactionid")
    )
    amount_sats = _wasabi_amount_sats(_wasabi_pick(sanitized, "amount", "amountSats"))
    is_coinjoin = _wasabi_bool(
        sanitized.get("islikelycoinjoin")
        if "islikelycoinjoin" in sanitized
        else sanitized.get("isLikelyCoinJoin")
        if "isLikelyCoinJoin" in sanitized
        else sanitized.get("isLikelyCoinjoin"),
        False,
    )
    label = str_or_none(sanitized.get("label") or sanitized.get("labels"))
    height = _wasabi_int(sanitized.get("height") or sanitized.get("blockHeight"), "height")
    direction = "outbound" if amount_sats < 0 or (amount_sats == 0 and is_coinjoin) else "inbound"
    kind = "coinjoin" if is_coinjoin else "withdrawal" if direction == "outbound" else "deposit"
    safe_raw = {
        "source": "wasabi_gethistory",
        "tx": txid,
        "height": height,
        "amount_sats": amount_sats,
        "label": label,
        "islikelycoinjoin": is_coinjoin,
        "privacy_hop": "coinjoin" if is_coinjoin else None,
    }
    safe_raw = {key: value for key, value in safe_raw.items() if value not in (None, "")}
    description = label or ("CoinJoin transaction" if is_coinjoin else "Wasabi transaction")
    return {
        "txid": txid,
        "occurred_at": sanitized.get("datetime") or sanitized.get("date") or sanitized.get("time"),
        "confirmed_at": sanitized.get("datetime") if height else None,
        "direction": direction,
        "asset": "BTC",
        "amount": _wasabi_btc_from_sats(abs(amount_sats)),
        "fee": Decimal("0"),
        "kind": kind,
        "description": description,
        "counterparty": None,
        "_wasabi_label": label,
        "_wasabi_islikelycoinjoin": is_coinjoin,
        "raw_json": json.dumps(json_ready(safe_raw), sort_keys=True),
    }


def _normalize_wasabi_coin_record(record: dict[str, Any]) -> dict[str, Any]:
    sanitized = {str(key): value for key, value in record.items() if key is not None}
    txid = _wasabi_safe_txid(
        sanitized.get("txid") or sanitized.get("tx") or sanitized.get("transactionId")
    )
    vout = _wasabi_int(
        sanitized.get("vout")
        if "vout" in sanitized
        else sanitized.get("index")
        if "index" in sanitized
        else sanitized.get("outputIndex"),
        "vout",
    )
    amount_sats = _wasabi_amount_sats(_wasabi_pick(sanitized, "amount", "amountSats"))
    confirmed = _wasabi_bool(sanitized.get("confirmed"), False)
    confirmations = _wasabi_int(sanitized.get("confirmations"), "confirmations", default=0)
    key_tail = _wasabi_key_path_tail(
        sanitized.get("keyPath") or sanitized.get("fullKeyPath") or sanitized.get("key_path")
    )
    spent_by = _wasabi_safe_txid(sanitized.get("spentBy") or sanitized.get("spent_by"))
    anon_history = (
        sanitized.get("anonHistory")
        or sanitized.get("anonymityHistory")
        or sanitized.get("anonScoreHistory")
        or sanitized.get("clusterHistory")
        or []
    )
    safe_raw = {
        "source": "wasabi_coin",
        "anonymityScore": sanitized.get("anonymityScore"),
        "excludedFromCoinjoin": sanitized.get("excludedFromCoinjoin"),
        "keyState": sanitized.get("keyState"),
        "has_key_path": key_tail["has_key_path"],
        "anon_history": _wasabi_sanitize_mapping(anon_history) if isinstance(anon_history, list) else [],
    }
    return {
        "txid": txid,
        "vout": vout,
        "amount_sats": amount_sats,
        "asset": "BTC",
        "chain": "bitcoin",
        "network": "mainnet",
        "confirmation_status": "confirmed" if confirmed else "mempool",
        "confirmations": confirmations,
        "block_height": _wasabi_int(sanitized.get("height") or sanitized.get("blockHeight"), "height"),
        "address": str_or_none(sanitized.get("address")),
        "address_label": str_or_none(sanitized.get("label")),
        "branch_label": key_tail["branch_label"],
        "branch_index": key_tail["branch_index"],
        "address_index": key_tail["address_index"],
        "anonymity_score": _wasabi_int(sanitized.get("anonymityScore"), "anonymityScore"),
        "spent_by": spent_by,
        "spent": bool(spent_by),
        "excluded_from_coinjoin": _wasabi_bool(sanitized.get("excludedFromCoinjoin"), False),
        "key_state": str_or_none(sanitized.get("keyState")),
        "anon_history": _wasabi_sanitize_mapping(anon_history) if isinstance(anon_history, list) else [],
        "raw": safe_raw,
    }


def _sanitize_wasabi_payment_in_coinjoin(record: dict[str, Any]) -> dict[str, Any]:
    states = record.get("state") or record.get("states") or []
    if isinstance(states, dict):
        states = [states]
    safe_states = []
    if isinstance(states, list):
        for state in states:
            if not isinstance(state, dict):
                continue
            safe_states.append(
                {
                    key: value
                    for key, value in {
                        "status": str_or_none(state.get("status")),
                        "round_id": str_or_none(state.get("roundId") or state.get("round_id")),
                        "payment_id": str_or_none(state.get("paymentId") or state.get("payment_id")),
                    }.items()
                    if value not in (None, "")
                }
            )
    return {
        key: value
        for key, value in {
            "id": str_or_none(record.get("id") or record.get("paymentId") or record.get("payment_id")),
            "round_id": str_or_none(record.get("roundId") or record.get("round_id")),
            "amount_sats": _wasabi_int(_wasabi_pick(record, "amount", "amountSats"), "payment amount"),
            "states": safe_states,
        }.items()
        if value not in (None, "", [])
    }


def load_wasabi_bundle(file_path: str) -> dict[str, Any]:
    """Load a sanitized Wasabi RPC/export bundle.

    The accepted shape is a JSON object containing RPC response bodies such as
    ``gethistory``, ``listcoins``/``listunspentcoins``, ``getwalletinfo``,
    ``listkeys``, ``listpaymentsincoinjoin``, and optional ``wallet_json``.
    Each section may be the raw ``result`` value or a JSON-RPC wrapper with a
    top-level ``result`` key. A bare list is treated as ``gethistory`` for
    backwards-compatible one-shot exports.
    """
    with open(file_path, "r", encoding="utf-8") as handle:
        payload = json.load(handle)
    return load_wasabi_bundle_payload(payload)


def load_wasabi_bundle_payload(payload: Any) -> dict[str, Any]:
    """Normalize a sanitized Wasabi RPC/export bundle payload."""
    if isinstance(payload, list):
        history_rows = [item for item in payload if isinstance(item, dict)]
        raw_payload: dict[str, Any] = {"gethistory": history_rows}
    elif isinstance(payload, dict):
        raw_payload = payload
        history_rows = _wasabi_list(raw_payload, "gethistory", "history", "transactions")
    else:
        raise AppError("Wasabi bundle must be a JSON object", code="validation")

    listcoins_present = _wasabi_section_present(raw_payload, "listcoins", "coins")
    listunspent_present = _wasabi_section_present(
        raw_payload, "listunspentcoins", "unspentcoins", "unspent"
    )
    listcoins = _wasabi_list(raw_payload, "listcoins", "coins")
    listunspent = _wasabi_list(raw_payload, "listunspentcoins", "unspentcoins", "unspent")
    coins_by_outpoint: dict[tuple[str | None, int | None], dict[str, Any]] = {}
    for source in (listcoins, listunspent):
        for coin in source:
            normalized = _normalize_wasabi_coin_record(coin)
            coins_by_outpoint[(normalized.get("txid"), normalized.get("vout"))] = normalized

    wallet_info = _wasabi_dict(raw_payload, "getwalletinfo", "walletinfo")
    wallet_json = _wasabi_dict(raw_payload, "wallet_json", "walletjson", "wallet")
    payments = [
        _sanitize_wasabi_payment_in_coinjoin(payment)
        for payment in _wasabi_list(raw_payload, "listpaymentsincoinjoin", "paymentsincoinjoin")
    ]
    listkeys = _wasabi_list(raw_payload, "listkeys", "keys")
    metadata = _wasabi_wallet_metadata(wallet_info, wallet_json)
    if listkeys:
        metadata["keyStateCounts"] = {}
        for key in listkeys:
            state = str_or_none(key.get("keyState"))
            if state:
                metadata["keyStateCounts"][state] = metadata["keyStateCounts"].get(state, 0) + 1
    if payments:
        metadata["paymentsInCoinJoin"] = payments

    records = [_normalize_wasabi_history_record(row) for row in history_rows]
    return {
        "records": records,
        "coins": list(coins_by_outpoint.values()),
        "metadata": metadata,
        "payments_in_coinjoin": payments,
        "coin_sections_present": listcoins_present or listunspent_present,
        "wallet_json_present": bool(wallet_json),
        "listkeys_count": len(listkeys),
    }


def load_wasabi_bundle_records(file_path: str) -> list[dict[str, Any]]:
    return load_wasabi_bundle(file_path)["records"]


def is_wasabi_format(input_format):
    return input_format == WASABI_BUNDLE_FORMAT


# -- BTCPay ------------------------------------------------------------------


def parse_btcpay_amount(amount_text, currency=None):
    """Parse a BTCPay amount cell like `"0.001 BTC"` into a signed `Decimal`."""
    if amount_text is None:
        raise AppError("BTCPay export is missing Amount")
    text = str(amount_text).strip()
    asset = str(currency or "BTC").strip().upper()
    suffixes = [asset, asset.lower(), asset.upper()]
    for suffix in suffixes:
        if suffix and text.endswith(suffix):
            text = text[: -len(suffix)].strip()
            break
    return dec(text)


def parse_btcpay_labels(value):
    """Normalize BTCPay labels from CSV, JSON export, or Greenfield shapes."""
    if value is None or value == "":
        return []
    labels = []
    if isinstance(value, dict):
        for key, item in value.items():
            text = item.get("text") if isinstance(item, dict) else item
            if text is None and key:
                text = key
            if text is None:
                continue
            label = str(text).strip()
            if label:
                labels.append(label)
        return labels
    if isinstance(value, list):
        for item in value:
            text = item.get("text") if isinstance(item, dict) else item
            if text is None:
                continue
            label = str(text).strip()
            if label:
                labels.append(label)
        return labels
    return [part.strip() for part in str(value).split(",") if part.strip()]


def normalize_btcpay_record(record):
    """Turn a raw BTCPay JSON/CSV row into the common import-record shape.

    Retains `_btcpay_comment` and `_btcpay_labels` as private keys so the
    coordinator's `apply_btcpay_metadata` step in
    `kassiber.core.imports` can apply them to the inserted transaction
    afterwards.
    """
    sanitized_record = {str(key): value for key, value in record.items() if key is not None}
    txid = sanitized_record.get("TransactionId") or sanitized_record.get("Transaction Id")
    timestamp = sanitized_record.get("Timestamp")
    currency = normalize_asset_code(sanitized_record.get("Currency") or "BTC")
    amount = parse_btcpay_amount(sanitized_record.get("Amount"), currency=currency)
    comment = sanitized_record.get("Comment")
    labels = parse_btcpay_labels(sanitized_record.get("Labels"))
    return {
        "txid": txid,
        "occurred_at": timestamp,
        "confirmed_at": sanitized_record.get("confirmed_at") or sanitized_record.get("ConfirmedAt"),
        "direction": "outbound" if amount < 0 else "inbound",
        "asset": currency,
        "amount": abs(amount),
        "fee": Decimal("0"),
        # BTCPay's Greenfield wallet-transactions API reports only the net wallet
        # balance change, with no per-tx fee field. For a send that net delta
        # already includes the miner fee, so flag it: the transfer-fee guard must
        # treat the out/in discrepancy as the fee, not an unrecognized outflow.
        "amount_includes_fee": amount < 0,
        "fiat_rate": None,
        "fiat_value": None,
        "kind": "withdrawal" if amount < 0 else "deposit",
        "description": comment or "Imported from BTCPay",
        "counterparty": None,
        "_btcpay_comment": comment,
        "_btcpay_labels": labels,
        "raw_json": json.dumps(json_ready(sanitized_record), sort_keys=True),
    }


def load_btcpay_export_records(file_path, input_format):
    """Load a BTCPay export file (JSON list or CSV) and normalize every row."""
    if input_format == "json":
        with open(file_path, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
        if not isinstance(payload, list):
            raise AppError("BTCPay JSON export must be a list of transaction objects")
        rows = payload
    elif input_format == "csv":
        with open(file_path, "r", encoding="utf-8", newline="") as handle:
            rows = list(csv.DictReader(handle))
    else:
        raise AppError(f"Unsupported BTCPay input format '{input_format}'")
    return [normalize_btcpay_record(row) for row in rows]


def is_btcpay_format(input_format):
    return input_format in {"btcpay_json", "btcpay_csv"}


# -- Phoenix -----------------------------------------------------------------


_PHOENIX_REQUIRED_COLUMNS = (
    "date",
    "id",
    "type",
    "amount_msat",
)

_PHOENIX_OUTBOUND_TYPES = {
    "lightning_sent",
    "swap_out",
    "legacy_swap_out",
    "channel_close",
    "liquidity_purchase",
    "fee_bumping",
}

_PHOENIX_INBOUND_TYPES = {
    "lightning_received",
    "swap_in",
    "legacy_swap_in",
    "legacy_pay_to_open",
}


def parse_phoenix_fiat_amount(amount_text):
    """Parse a Phoenix amount_fiat cell like `"22.9998 USD"` into `(Decimal, currency)`."""
    if amount_text is None:
        return None, None
    text = str(amount_text).strip()
    if not text:
        return None, None
    parts = text.split()
    if len(parts) == 1:
        return dec(parts[0]), None
    value = dec(parts[0])
    currency = normalize_asset_code(parts[1])
    return value, currency


def _normalize_payment_hash(value):
    """Normalize a Lightning payment-hash string to 64-char lowercase hex.

    Returns ``None`` when the value is empty, non-hex, or not exactly 32
    bytes. The matcher only trusts values that round-trip cleanly.
    """
    text = str_or_none(value)
    if not text:
        return None
    text = text.strip().lower()
    if len(text) != 64:
        return None
    try:
        bytes.fromhex(text)
    except ValueError:
        return None
    return text


def _payment_hash_from_preimage(value):
    """Derive a Lightning payment hash from a 32-byte hex preimage."""
    text = str_or_none(value)
    if not text:
        return None
    text = text.strip().lower()
    if len(text) != 64:
        return None
    try:
        preimage = bytes.fromhex(text)
    except ValueError:
        return None
    return hashlib.sha256(preimage).hexdigest()


def normalize_phoenix_record(record):
    """Turn a Phoenix CSV row into the common import-record shape.

    Phoenix ships amounts as signed INTEGER msat, fees split across
    `mining_fee_sat` (×1000 to msat) + `service_fee_msat`, and fiat as a
    signed "<value> <CCY>" string. We take the absolute value everywhere
    (direction is captured separately) and derive `fiat_rate` from
    `fiat_value / amount_btc` since Phoenix does not export the rate.

    Lightning rows expose a ``payment_hash`` column that we promote to
    the canonical ``payment_hash`` field consumed by
    ``kassiber.core.imports.normalize_import_record`` so the matcher can
    pair the LN leg of a submarine swap with its on-chain counterpart
    deterministically.

    Private keys `_phoenix_type` and `_phoenix_description` feed
    `apply_phoenix_metadata` in `kassiber.core.imports`.
    """
    sanitized = {str(key): value for key, value in record.items() if key is not None}
    for column in _PHOENIX_REQUIRED_COLUMNS:
        if column not in sanitized:
            raise AppError(f"Phoenix CSV is missing required column '{column}'")
    phoenix_type = str(sanitized.get("type") or "").strip() or "unknown"
    amount_msat_raw = str(sanitized.get("amount_msat") or "0").strip() or "0"
    try:
        amount_msat_signed = int(amount_msat_raw)
    except ValueError as exc:
        raise AppError(f"Invalid Phoenix amount_msat '{amount_msat_raw}'") from exc
    if amount_msat_signed < 0:
        direction = "outbound"
    elif amount_msat_signed > 0:
        direction = "inbound"
    elif phoenix_type in _PHOENIX_OUTBOUND_TYPES:
        direction = "outbound"
    elif phoenix_type in _PHOENIX_INBOUND_TYPES:
        direction = "inbound"
    else:
        direction = "outbound"
    amount_btc = msat_to_btc(abs(amount_msat_signed))
    mining_fee_sat_raw = str(sanitized.get("mining_fee_sat") or "0").strip() or "0"
    service_fee_msat_raw = str(sanitized.get("service_fee_msat") or "0").strip() or "0"
    try:
        mining_fee_msat = int(mining_fee_sat_raw) * 1000
    except ValueError as exc:
        raise AppError(f"Invalid Phoenix mining_fee_sat '{mining_fee_sat_raw}'") from exc
    try:
        service_fee_msat = int(service_fee_msat_raw)
    except ValueError as exc:
        raise AppError(f"Invalid Phoenix service_fee_msat '{service_fee_msat_raw}'") from exc
    fee_btc = msat_to_btc(mining_fee_msat + service_fee_msat)
    fiat_value_signed, _ = parse_phoenix_fiat_amount(sanitized.get("amount_fiat"))
    fiat_value = abs(fiat_value_signed) if fiat_value_signed is not None else None
    fiat_rate = None
    if fiat_value is not None and amount_btc > 0:
        fiat_rate = fiat_value / amount_btc
    description = str_or_none(sanitized.get("description"))
    counterparty = str_or_none(sanitized.get("destination"))
    payment_hash = _normalize_payment_hash(sanitized.get("payment_hash"))
    return {
        "txid": sanitized.get("id"),
        "occurred_at": sanitized.get("date"),
        "direction": direction,
        "asset": "BTC",
        "amount": amount_btc,
        "fee": fee_btc,
        "fiat_rate": fiat_rate,
        "fiat_value": fiat_value,
        "kind": phoenix_type,
        "description": description,
        "counterparty": counterparty,
        "payment_hash": payment_hash,
        "payment_hash_source": "importer" if payment_hash else None,
        "_phoenix_type": phoenix_type,
        "_phoenix_description": description,
        "raw_json": json.dumps(json_ready(sanitized), sort_keys=True),
    }


def load_phoenix_csv_records(file_path):
    """Load a Phoenix CSV export and normalize every row.

    Validates the 4 required columns up-front so the user gets one
    error envelope rather than a cascade of row-level failures.
    """
    with open(file_path, "r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if not rows:
        return []
    header = rows[0].keys()
    missing = [column for column in _PHOENIX_REQUIRED_COLUMNS if column not in header]
    if missing:
        raise AppError(
            "Phoenix CSV is missing required columns: " + ", ".join(missing)
        )
    return [normalize_phoenix_record(row) for row in rows]


def is_phoenix_format(input_format):
    return input_format == "phoenix_csv"


# -- River -------------------------------------------------------------------


_RIVER_REQUIRED_COLUMNS = (
    "Date",
    "Sent Amount",
    "Sent Currency",
    "Received Amount",
    "Received Currency",
)


def _normalized_column_key(value):
    return " ".join(str(value).replace("\xa0", " ").strip().split()).casefold()


def _casefold_record(record):
    output = {}
    for key, value in record.items():
        if key is None:
            continue
        output[_normalized_column_key(key)] = value
    return output


def _get_cell(record, *names):
    folded = _casefold_record(record)
    for name in names:
        value = folded.get(_normalized_column_key(name))
        if value not in (None, ""):
            return value
    return None


def _clean_decimal_text(value):
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text:
        return None
    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1].strip()
    for symbol in ("$", "€", "£", "CHF", "USD", "EUR", "GBP", "BTC", "XBT", "LBTC"):
        text = text.replace(symbol, "")
    text = text.replace(",", "").replace(" ", "").strip()
    if not text:
        return None
    if negative and not text.startswith("-"):
        text = "-" + text
    return text


def _decimal_cell(value):
    text = _clean_decimal_text(value)
    return dec(text) if text is not None else None


def _currency_cell(value):
    currency = str_or_none(value)
    return normalize_asset_code(currency) if currency else None


def _river_decimal(value):
    return _decimal_cell(value)


def _river_currency(value):
    return _currency_cell(value)


_BITCOIN_FAMILY_ASSETS = {"BTC", "XBT", "LBTC"}
_FIAT_CURRENCIES = {
    "AED",
    "ARS",
    "AUD",
    "BRL",
    "CAD",
    "CHF",
    "CLP",
    "CNY",
    "CZK",
    "DKK",
    "EUR",
    "GBP",
    "HKD",
    "HUF",
    "ILS",
    "INR",
    "JPY",
    "KRW",
    "MXN",
    "NOK",
    "NZD",
    "PLN",
    "RON",
    "SEK",
    "SGD",
    "THB",
    "TRY",
    "USD",
    "ZAR",
}


def _bitcoin_family_asset(value):
    asset = _currency_cell(value)
    if asset == "XBT":
        return "BTC"
    if asset in _BITCOIN_FAMILY_ASSETS:
        return asset
    return None


def _is_fiat_currency(value):
    return _currency_cell(value) in _FIAT_CURRENCIES


# -- Bull Bitcoin ------------------------------------------------------------


_BULLBITCOIN_REQUIRED_COLUMNS = (
    "ORDER_STATUS",
    "PAYIN_AMOUNT",
    "PAYIN_CURRENCY",
    "PAYOUT_AMOUNT",
    "PAYOUT_CURRENCY",
    "COMPLETED_AT (UTC)",
    "TRANSACTION_ID",
)

_BULLBITCOIN_CRYPTO_CURRENCIES = {"BTC", "XBT", "LBTC"}


def _bullbitcoin_completed(record):
    for column in ("ORDER_STATUS", "PAYIN_STATUS", "PAYOUT_STATUS"):
        value = str_or_none(_get_cell(record, column))
        if value and value.strip().casefold() != "completed":
            return False
    return True


def normalize_bullbitcoin_record(record):
    """Turn one Bull Bitcoin order-export row into the common import shape."""
    sanitized = {
        str(key).strip(): value for key, value in record.items() if key is not None
    }
    if not _bullbitcoin_completed(sanitized):
        return None

    payin_currency = _currency_cell(_get_cell(sanitized, "PAYIN_CURRENCY"))
    payout_currency = _currency_cell(_get_cell(sanitized, "PAYOUT_CURRENCY"))
    payin_amount = _decimal_cell(_get_cell(sanitized, "PAYIN_AMOUNT"))
    payout_amount = _decimal_cell(_get_cell(sanitized, "PAYOUT_AMOUNT"))
    if payin_amount is None or payout_amount is None:
        raise AppError("Bull Bitcoin CSV has a completed row with an empty amount")

    payin_is_crypto = payin_currency in _BULLBITCOIN_CRYPTO_CURRENCIES
    payout_is_crypto = payout_currency in _BULLBITCOIN_CRYPTO_CURRENCIES
    if payin_is_crypto and not payout_is_crypto:
        direction = "outbound"
        asset = "BTC" if payin_currency == "XBT" else payin_currency
        amount = abs(payin_amount)
        fiat_value = abs(payout_amount)
        fiat_currency = payout_currency
        kind = "sell"
    elif payout_is_crypto and not payin_is_crypto:
        direction = "inbound"
        asset = "BTC" if payout_currency == "XBT" else payout_currency
        amount = abs(payout_amount)
        fiat_value = abs(payin_amount)
        fiat_currency = payin_currency
        kind = "buy"
    else:
        return None

    txid = str_or_none(_get_cell(sanitized, "TRANSACTION_ID"))
    if not txid:
        return None
    exchange_rate = _decimal_cell(_get_cell(sanitized, "EXCHANGE_RATE_AMOUNT"))
    fiat_rate = exchange_rate if exchange_rate is not None else fiat_value / amount
    order_ref = (
        str_or_none(_get_cell(sanitized, "ORDER_ID"))
        or str_or_none(_get_cell(sanitized, "ORDER_NUMBER"))
        or txid
    )
    occurred_at = (
        _get_cell(sanitized, "COMPLETED_AT (UTC)")
        or _get_cell(sanitized, "SENT_AT (UTC)")
        or _get_cell(sanitized, "CREATED_AT (UTC)")
    )
    payin_method = str_or_none(_get_cell(sanitized, "PAYIN_METHOD"))
    payout_method = str_or_none(_get_cell(sanitized, "PAYOUT_METHOD"))
    method_text = " to ".join(part for part in (payin_method, payout_method) if part)
    description = f"Bull Bitcoin {kind}"
    if method_text:
        description = f"{description} - {method_text}"
    return {
        "txid": txid,
        "occurred_at": occurred_at,
        "direction": direction,
        "asset": asset,
        "amount": amount,
        "fee": Decimal("0"),
        "fiat_rate": fiat_rate,
        "fiat_value": fiat_value,
        "fiat_currency": fiat_currency,
        "pricing_source_kind": "exchange_execution",
        "pricing_provider": "Bull Bitcoin",
        "pricing_pair": f"{asset}-{fiat_currency}" if fiat_currency else None,
        "pricing_method": "bullbitcoin_csv",
        "pricing_external_ref": order_ref,
        "pricing_quality": "exact",
        "kind": kind,
        "description": description,
        "counterparty": "Bull Bitcoin",
        "raw_json": json.dumps(json_ready(sanitized), sort_keys=True),
    }


def load_bullbitcoin_csv_records(file_path):
    """Load Bull Bitcoin order-export CSV rows."""
    with open(file_path, "r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if not rows:
        return []
    header = {_normalized_column_key(column) for column in rows[0].keys()}
    missing = [
        column for column in _BULLBITCOIN_REQUIRED_COLUMNS if _normalized_column_key(column) not in header
    ]
    if missing:
        raise AppError("Bull Bitcoin CSV is missing required columns: " + ", ".join(missing))
    normalized = []
    for row in rows:
        record = normalize_bullbitcoin_record(row)
        if record is not None:
            normalized.append(record)
    return normalized


def is_bullbitcoin_format(input_format):
    return input_format == "bullbitcoin_csv"


# -- Bull Bitcoin wallet export ---------------------------------------------


BULLBITCOIN_WALLET_CSV_FORMAT = "bullbitcoin_wallet_csv"

_BULLBITCOIN_WALLET_REQUIRED_COLUMNS = (
    "date",
    "type",
    "direction",
    "amount_sats",
    "amount_btc",
    "fee_sats",
    "status",
    "txid",
    "network",
    "address",
    "swap_id",
    "preimage",
    "total_swap_fees_sats",
    "send_network",
    "receive_network",
    "send_txid",
    "receive_txid",
)

_BULLBITCOIN_WALLET_SKIPPED_STATUSES = {"failed", "expired", "refunded"}
_BULLBITCOIN_WALLET_INBOUND_DIRECTIONS = {"incoming", "inbound", "received", "receive"}
_BULLBITCOIN_WALLET_OUTBOUND_DIRECTIONS = {"outgoing", "outbound", "sent", "send"}


def _bullbitcoin_wallet_sats(record, column, *, default=None):
    value = _decimal_cell(_get_cell(record, column))
    if value is None:
        if default is not None:
            return default
        raise AppError(f"Bull Bitcoin wallet CSV has a row with empty {column}")
    integral = value.to_integral_value()
    if value != integral:
        raise AppError(f"Bull Bitcoin wallet CSV has non-integer {column}: {value}")
    return int(integral)


def _bullbitcoin_wallet_btc_from_sats(sats):
    return Decimal(sats) / Decimal("100000000")


def _bullbitcoin_wallet_direction(value):
    direction = str(value or "").strip().casefold()
    if direction in _BULLBITCOIN_WALLET_INBOUND_DIRECTIONS:
        return "inbound"
    if direction in _BULLBITCOIN_WALLET_OUTBOUND_DIRECTIONS:
        return "outbound"
    if direction == "self":
        return None
    raise AppError(f"Unsupported Bull Bitcoin wallet direction '{value}'")


def _bullbitcoin_wallet_network(record, bull_type, direction):
    network = str_or_none(_get_cell(record, "network"))
    if not network and bull_type == "chain_swap":
        network = str_or_none(
            _get_cell(
                record,
                "receive_network" if direction == "inbound" else "send_network",
            )
        )
    if not network and bull_type == "liquid":
        network = "liquid"
    if not network and bull_type in {"onchain", "payjoin_send", "payjoin_receive"}:
        network = "bitcoin"
    if not network and bull_type in {"lightning_send", "lightning_receive"}:
        network = "lightning"
    normalized = str(network or "").strip().casefold()
    if normalized in {"bitcoin", "btc", "onchain"}:
        return "bitcoin"
    if normalized in {"liquid", "lbtc", "liquidv1"}:
        return "liquid"
    if normalized in {"lightning", "ln"}:
        return "lightning"
    raise AppError(f"Unsupported Bull Bitcoin wallet network '{network}'")


def _bullbitcoin_wallet_asset(network, bull_type):
    if network == "liquid" or bull_type == "liquid":
        return "LBTC"
    return "BTC"


def _bullbitcoin_wallet_external_id(sanitized, index):
    txid = str_or_none(_get_cell(sanitized, "txid"))
    if txid:
        return txid
    for column in ("receive_txid", "send_txid", "swap_id"):
        value = str_or_none(_get_cell(sanitized, column))
        if value:
            return value
    return (
        "bullbitcoin-wallet:"
        f"{_get_cell(sanitized, 'date') or index}:"
        f"{_get_cell(sanitized, 'type') or 'transaction'}:"
        f"{_get_cell(sanitized, 'direction') or ''}:"
        f"{_get_cell(sanitized, 'amount_sats') or ''}"
    )


def _bullbitcoin_wallet_payment_hash(bull_type, external_id, preimage):
    if bull_type not in {"lightning_send", "lightning_receive"}:
        return None
    return _payment_hash_from_preimage(preimage) or _normalize_payment_hash(external_id)


def normalize_bullbitcoin_wallet_record(record, index=0):
    """Turn one Bull Bitcoin unified wallet-history row into the common shape."""
    sanitized = {str(key).strip(): value for key, value in record.items() if key is not None}
    bull_type = str(_get_cell(sanitized, "type") or "").strip().casefold()
    status = str(_get_cell(sanitized, "status") or "").strip().casefold()
    if status in _BULLBITCOIN_WALLET_SKIPPED_STATUSES:
        return None
    direction = _bullbitcoin_wallet_direction(_get_cell(sanitized, "direction"))
    if direction is None:
        return None
    amount_sats = abs(_bullbitcoin_wallet_sats(sanitized, "amount_sats"))
    fee_sats = abs(_bullbitcoin_wallet_sats(sanitized, "fee_sats", default=0))
    network = _bullbitcoin_wallet_network(sanitized, bull_type, direction)
    asset = _bullbitcoin_wallet_asset(network, bull_type)
    external_id = _bullbitcoin_wallet_external_id(sanitized, index)
    payment_hash = _bullbitcoin_wallet_payment_hash(
        bull_type,
        external_id,
        _get_cell(sanitized, "preimage"),
    )
    raw_payload = dict(sanitized)
    preimage_redacted = False
    for key in list(raw_payload.keys()):
        if _normalized_column_key(key) == "preimage" and str_or_none(raw_payload.get(key)):
            raw_payload[key] = "[redacted]"
            preimage_redacted = True
    if preimage_redacted:
        raw_payload["preimage_redacted"] = True
    raw_payload["source"] = BULLBITCOIN_WALLET_CSV_FORMAT
    raw_payload["normalized_network"] = network

    record_out = {
        "txid": external_id,
        "occurred_at": _get_cell(sanitized, "date"),
        "confirmed_at": _get_cell(sanitized, "date")
        if status in {"confirmed", "completed"}
        else None,
        "direction": direction,
        "asset": asset,
        "amount": _bullbitcoin_wallet_btc_from_sats(amount_sats),
        "fee": _bullbitcoin_wallet_btc_from_sats(fee_sats),
        "fiat_rate": None,
        "fiat_value": None,
        "kind": bull_type or "transaction",
        "description": f"Bull Bitcoin wallet {bull_type or 'transaction'}",
        "counterparty": None,
        "payment_hash": payment_hash,
        "payment_hash_source": "importer" if payment_hash else None,
        "_bullbitcoin_wallet_network": network,
        # Bull reports a per-transaction fee on receive/swap rows, but a wallet
        # that synced the same on-chain transaction from a descriptor backend
        # stores a receive fee of 0 (the recipient pays none). Let enrichment
        # match those rows by txid/amount even when the fee differs; merges only
        # attach metadata and never overwrite the stored fee.
        "_match_existing_ignore_fee": True,
        "raw_json": json.dumps(json_ready(raw_payload), sort_keys=True),
    }
    if bull_type in {"payjoin_send", "payjoin_receive"}:
        record_out["privacy_boundary"] = "payjoin"
    return record_out


def load_bullbitcoin_wallet_csv_records(file_path):
    """Load Bull Bitcoin unified wallet-history CSV rows."""
    with open(file_path, "r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if not rows:
        return []
    header = {_normalized_column_key(column) for column in rows[0].keys()}
    missing = [
        column
        for column in _BULLBITCOIN_WALLET_REQUIRED_COLUMNS
        if _normalized_column_key(column) not in header
    ]
    if missing:
        raise AppError(
            "Bull Bitcoin wallet CSV is missing required columns: "
            + ", ".join(missing)
        )
    normalized = []
    for index, row in enumerate(rows, start=1):
        record = normalize_bullbitcoin_wallet_record(row, index=index)
        if record is not None:
            normalized.append(record)
    return normalized


def is_bullbitcoin_wallet_format(input_format):
    return input_format == BULLBITCOIN_WALLET_CSV_FORMAT


def bullbitcoin_wallet_record_network(record):
    network = str_or_none(record.get("_bullbitcoin_wallet_network"))
    if network:
        return network
    raw_json = record.get("raw_json")
    if not raw_json:
        return None
    try:
        raw_payload = json.loads(raw_json)
    except (TypeError, ValueError):
        return None
    return str_or_none(raw_payload.get("normalized_network"))


# -- Coinfinity --------------------------------------------------------------


_COINFINITY_REQUIRED_COLUMNS = (
    "Order ID",
    "Type",
    "Date",
    "Amount EUR",
    "Amount Crypto",
    "Crypto",
    "Rate EUR",
    "Mining Fee Crypto",
    "Total Fee EUR",
    "Transaction",
)

_COINFINITY_CRYPTO_CURRENCIES = {"BTC", "XBT"}


def _coinfinity_user_direction(order_type):
    """Map Coinfinity's broker-side type to the user's BTC movement."""
    value = str(order_type or "").strip().casefold()
    if value == "sell":
        return "inbound", "buy"
    if value == "buy":
        return "outbound", "sell"
    return None, None


def normalize_coinfinity_record(record, index=0):
    """Turn one Coinfinity order-export row into the common import shape."""
    sanitized = {str(key).strip(): value for key, value in record.items() if key is not None}
    direction, kind = _coinfinity_user_direction(_get_cell(sanitized, "Type"))
    if direction is None or kind is None:
        return None

    currency = _currency_cell(_get_cell(sanitized, "Crypto"))
    if currency not in _COINFINITY_CRYPTO_CURRENCIES:
        return None
    asset = "BTC" if currency == "XBT" else currency
    amount = _decimal_cell(_get_cell(sanitized, "Amount Crypto"))
    fiat_amount = _decimal_cell(_get_cell(sanitized, "Amount EUR"))
    fee_fiat = _decimal_cell(_get_cell(sanitized, "Total Fee EUR")) or Decimal("0")
    if amount is None or fiat_amount is None:
        raise AppError("Coinfinity CSV has a BTC/EUR row with an empty amount")

    if direction == "inbound":
        fiat_value = abs(fiat_amount) + abs(fee_fiat)
        btc_fee = Decimal("0")
    else:
        fiat_value = max(Decimal("0"), abs(fiat_amount) - abs(fee_fiat))
        btc_fee = _decimal_cell(_get_cell(sanitized, "Mining Fee Crypto")) or Decimal(
            "0"
        )

    fiat_rate = _decimal_cell(_get_cell(sanitized, "Rate EUR"))
    occurred_at = _get_cell(sanitized, "Date")
    order_ref = str_or_none(_get_cell(sanitized, "Order ID")) or (
        f"coinfinity:{occurred_at}:{direction}:{asset}:{amount}:{index}"
    )
    transaction_ref = str_or_none(_get_cell(sanitized, "Transaction"))
    lightning_ref = str_or_none(_get_cell(sanitized, "LN Invoice"))
    external_id = transaction_ref or lightning_ref or f"coinfinity:{order_ref}"
    transaction_type = str_or_none(_get_cell(sanitized, "Transaction type"))
    description = f"Coinfinity {kind}"
    if transaction_type:
        description = f"{description} - {transaction_type}"
    payload = {
        "txid": external_id,
        "occurred_at": occurred_at,
        "direction": direction,
        "asset": asset,
        "amount": abs(amount),
        "fee": abs(btc_fee),
        "fiat_rate": (
            fiat_rate
            if fiat_rate is not None
            else (fiat_value / abs(amount) if amount else None)
        ),
        "fiat_value": fiat_value,
        "fiat_currency": "EUR",
        "pricing_source_kind": "exchange_execution",
        "pricing_provider": "Coinfinity",
        "pricing_pair": f"{asset}-EUR",
        "pricing_timestamp": occurred_at,
        "pricing_method": "coinfinity_csv",
        "pricing_external_ref": order_ref,
        "pricing_quality": "exact",
        "kind": kind,
        "description": description,
        "counterparty": "Coinfinity",
        "raw_json": json.dumps(json_ready(sanitized), sort_keys=True),
    }
    if not transaction_ref and not lightning_ref:
        payload["_exchange_evidence_match_by_economics"] = True
        payload["_exchange_evidence_match_time_tolerance_seconds"] = 172800
    return payload


def load_coinfinity_csv_records(file_path):
    """Load Coinfinity order-export CSV rows."""
    with open(file_path, "r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if not rows:
        return []
    header = {_normalized_column_key(column) for column in rows[0].keys()}
    missing = [
        column
        for column in _COINFINITY_REQUIRED_COLUMNS
        if _normalized_column_key(column) not in header
    ]
    if missing:
        raise AppError("Coinfinity CSV is missing required columns: " + ", ".join(missing))
    normalized = []
    for index, row in enumerate(rows, start=1):
        record = normalize_coinfinity_record(row, index=index)
        if record is not None:
            normalized.append(record)
    return normalized


def is_coinfinity_format(input_format):
    return input_format == "coinfinity_csv"


# -- Pocket Bitcoin ----------------------------------------------------------


_POCKETBITCOIN_REQUIRED_COLUMNS = (
    "type",
    "date",
    "reference",
    "price.currency",
    "price.amount",
    "cost.currency",
    "cost.amount",
    "fee.currency",
    "fee.amount",
    "value.currency",
    "value.amount",
)

_POCKETBITCOIN_CRYPTO_CURRENCIES = {"BTC", "XBT", "LBTC"}


def _pocketbitcoin_row_type(record):
    return str(_get_cell(record, "type") or "").strip().casefold()


def _pocketbitcoin_crypto_asset(currency):
    if currency in _POCKETBITCOIN_CRYPTO_CURRENCIES:
        return "BTC" if currency == "XBT" else currency
    return None


def _pocketbitcoin_reference(record, fallback):
    return str_or_none(_get_cell(record, "reference")) or fallback


def _pocketbitcoin_withdrawal_key(withdrawal):
    currency = _currency_cell(_get_cell(withdrawal, "value.currency"))
    asset = _pocketbitcoin_crypto_asset(currency)
    amount = _decimal_cell(_get_cell(withdrawal, "value.amount"))
    fee = _decimal_cell(_get_cell(withdrawal, "fee.amount"))
    fee_currency = _currency_cell(_get_cell(withdrawal, "fee.currency"))
    if asset is None or amount is None or fee is None or fee_currency != currency:
        return None
    return asset, abs(amount) + abs(fee)


def _pocketbitcoin_withdrawal_pairs(rows):
    """Match Pocket exchange gross BTC amounts to separate withdrawal rows."""
    withdrawals_by_key = {}
    for row in rows:
        if _pocketbitcoin_row_type(row) != "withdrawal":
            continue
        key = _pocketbitcoin_withdrawal_key(row)
        if key is None:
            continue
        withdrawals_by_key.setdefault(key, []).append(row)
    for bucket in withdrawals_by_key.values():
        bucket.sort(key=lambda item: str(_get_cell(item, "date") or ""))

    pairs = {}
    for row in rows:
        if _pocketbitcoin_row_type(row) != "exchange":
            continue
        value_currency = _currency_cell(_get_cell(row, "value.currency"))
        asset = _pocketbitcoin_crypto_asset(value_currency)
        gross_amount = _decimal_cell(_get_cell(row, "value.amount"))
        if asset is None or gross_amount is None:
            continue
        candidates = withdrawals_by_key.get((asset, abs(gross_amount)), [])
        if not candidates:
            continue
        exchange_date = str(_get_cell(row, "date") or "")
        chosen_index = None
        for index, candidate in enumerate(candidates):
            withdrawal_date = str(_get_cell(candidate, "date") or "")
            if not exchange_date or not withdrawal_date or withdrawal_date >= exchange_date:
                chosen_index = index
                break
        if chosen_index is None:
            chosen_index = 0
        pairs[id(row)] = candidates.pop(chosen_index)
    return pairs


def normalize_pocketbitcoin_record(record, withdrawal=None, index=0):
    """Turn one Pocket Bitcoin exchange row into the common import shape."""
    sanitized = {str(key).strip(): value for key, value in record.items() if key is not None}
    if _pocketbitcoin_row_type(sanitized) != "exchange":
        return None

    cost_currency = _currency_cell(_get_cell(sanitized, "cost.currency"))
    value_currency = _currency_cell(_get_cell(sanitized, "value.currency"))
    fee_currency = _currency_cell(_get_cell(sanitized, "fee.currency"))
    cost_amount = _decimal_cell(_get_cell(sanitized, "cost.amount"))
    value_amount = _decimal_cell(_get_cell(sanitized, "value.amount"))
    fee_amount = _decimal_cell(_get_cell(sanitized, "fee.amount"))
    if cost_amount is None or value_amount is None:
        raise AppError("Pocket Bitcoin CSV has an exchange row with an empty amount")

    cost_asset = _pocketbitcoin_crypto_asset(cost_currency)
    value_asset = _pocketbitcoin_crypto_asset(value_currency)
    if value_asset and not cost_asset:
        direction = "inbound"
        asset = value_asset
        amount = abs(value_amount)
        fiat_value = abs(cost_amount)
        fiat_currency = cost_currency
        kind = "buy"
        if fee_currency == fiat_currency and fee_amount is not None:
            fiat_value += abs(fee_amount)
    elif cost_asset and not value_asset:
        direction = "outbound"
        asset = cost_asset
        amount = abs(cost_amount)
        fiat_value = abs(value_amount)
        fiat_currency = value_currency
        kind = "sell"
        if fee_currency == fiat_currency and fee_amount is not None:
            fiat_value = max(Decimal("0"), fiat_value - abs(fee_amount))
    else:
        return None

    btc_fee = Decimal("0")
    occurred_at = _get_cell(sanitized, "date")
    raw_payload = dict(sanitized)
    if withdrawal is not None:
        withdrawal_currency = _currency_cell(_get_cell(withdrawal, "value.currency"))
        withdrawal_asset = _pocketbitcoin_crypto_asset(withdrawal_currency)
        withdrawal_amount = _decimal_cell(_get_cell(withdrawal, "value.amount"))
        withdrawal_fee = _decimal_cell(_get_cell(withdrawal, "fee.amount"))
        withdrawal_fee_currency = _currency_cell(_get_cell(withdrawal, "fee.currency"))
        if (
            direction == "inbound"
            and withdrawal_asset == asset
            and withdrawal_amount is not None
        ):
            amount = abs(withdrawal_amount)
            occurred_at = _get_cell(withdrawal, "date") or occurred_at
            if withdrawal_fee_currency == withdrawal_currency and withdrawal_fee is not None:
                btc_fee = abs(withdrawal_fee)
        raw_payload["_pocketbitcoin_withdrawal"] = {
            str(key).strip(): value for key, value in withdrawal.items() if key is not None
        }

    price_amount = _decimal_cell(_get_cell(sanitized, "price.amount"))
    fiat_rate = price_amount if price_amount is not None else (fiat_value / amount if amount else None)
    reference = _pocketbitcoin_reference(
        sanitized,
        f"pocketbitcoin:{occurred_at}:{direction}:{asset}:{amount}:{index}",
    )
    description = f"Pocket Bitcoin {kind}"
    return {
        "txid": f"pocketbitcoin:{reference}",
        "occurred_at": occurred_at,
        "direction": direction,
        "asset": asset,
        "amount": amount,
        "fee": btc_fee,
        "fiat_rate": fiat_rate,
        "fiat_value": fiat_value,
        "fiat_currency": fiat_currency,
        "pricing_source_kind": "exchange_execution",
        "pricing_provider": "Pocket Bitcoin",
        "pricing_pair": f"{asset}-{fiat_currency}" if fiat_currency else None,
        "pricing_timestamp": _get_cell(sanitized, "date"),
        "pricing_method": "pocketbitcoin_csv",
        "pricing_external_ref": reference,
        "pricing_quality": "exact",
        "kind": kind,
        "description": description,
        "counterparty": "Pocket Bitcoin",
        "_exchange_evidence_match_by_economics": True,
        "_exchange_evidence_match_time_tolerance_seconds": 172800,
        "raw_json": json.dumps(json_ready(raw_payload), sort_keys=True),
    }


def load_pocketbitcoin_csv_records(file_path):
    """Load Pocket Bitcoin account CSV rows."""
    with open(file_path, "r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if not rows:
        return []
    header = {_normalized_column_key(column) for column in rows[0].keys()}
    missing = [
        column
        for column in _POCKETBITCOIN_REQUIRED_COLUMNS
        if _normalized_column_key(column) not in header
    ]
    if missing:
        raise AppError("Pocket Bitcoin CSV is missing required columns: " + ", ".join(missing))
    withdrawals = _pocketbitcoin_withdrawal_pairs(rows)
    normalized = []
    for index, row in enumerate(rows, start=1):
        record = normalize_pocketbitcoin_record(row, withdrawals.get(id(row)), index=index)
        if record is not None:
            normalized.append(record)
    return normalized


def is_pocketbitcoin_format(input_format):
    return input_format == "pocketbitcoin_csv"


# -- 21bitcoin ---------------------------------------------------------------


_TWENTYONEBITCOIN_REQUIRED_COLUMNS = (
    "id",
    "transaction_date",
    "buy_asset",
    "buy_amount",
    "sell_asset",
    "sell_amount",
    "fee_asset",
    "fee_amount",
    "transaction_type",
)

_TWENTYONEBITCOIN_CRYPTO_CURRENCIES = {"BTC", "XBT", "LBTC"}


def _twentyonebitcoin_asset(value):
    asset = _currency_cell(value)
    return "BTC" if asset == "XBT" else asset


def _twentyonebitcoin_datetime(value):
    text = str_or_none(value)
    if not text:
        return None
    text = text.strip()
    date_part, _, time_part = text.partition(" ")
    date_bits = date_part.split(".")
    if len(date_bits) != 3:
        return text
    day, month, year_text = date_bits
    year_text = year_text.strip()
    time_bits = (time_part.strip() or "00:00:00").split(":")
    if len(time_bits) != 3:
        return text
    try:
        if len(year_text) == 3 and year_text.startswith("2"):
            year = 2000 + int(year_text[-2:])
        elif len(year_text) == 2:
            year = 2000 + int(year_text)
        else:
            year = int(year_text)
        second = time_bits[2].split(".", 1)[0]
        dt = datetime(
            year,
            int(month),
            int(day),
            int(time_bits[0]),
            int(time_bits[1]),
            int(second),
            tzinfo=timezone.utc,
        )
    except ValueError:
        return text
    return dt.replace(microsecond=0).isoformat().replace("+00:00", "Z")


def normalize_twentyonebitcoin_record(record):
    """Turn one 21bitcoin transaction-export row into the common import shape."""
    sanitized = {str(key).strip(): value for key, value in record.items() if key is not None}
    row_id = str_or_none(_get_cell(sanitized, "id"))
    if not row_id:
        raise AppError("21bitcoin CSV has a BTC-side row without an id")

    buy_currency = _twentyonebitcoin_asset(_get_cell(sanitized, "buy_asset"))
    sell_currency = _twentyonebitcoin_asset(_get_cell(sanitized, "sell_asset"))
    fee_currency = _twentyonebitcoin_asset(_get_cell(sanitized, "fee_asset"))
    buy_amount = _decimal_cell(_get_cell(sanitized, "buy_amount"))
    sell_amount = _decimal_cell(_get_cell(sanitized, "sell_amount"))
    fee_amount = abs(_decimal_cell(_get_cell(sanitized, "fee_amount")) or Decimal("0"))

    buy_is_crypto = buy_currency in _TWENTYONEBITCOIN_CRYPTO_CURRENCIES
    sell_is_crypto = sell_currency in _TWENTYONEBITCOIN_CRYPTO_CURRENCIES
    fee_is_crypto = fee_currency in _TWENTYONEBITCOIN_CRYPTO_CURRENCIES
    fee_is_fiat = fee_currency is not None and not fee_is_crypto
    transaction_type = str_or_none(_get_cell(sanitized, "transaction_type")) or "transaction"
    transaction_type_key = transaction_type.strip().lower()
    is_l1_withdrawal = transaction_type_key == "withdrawal"

    fiat_currency = None
    fiat_value = None
    fiat_rate = None
    pricing_source_kind = None
    pricing_quality = None
    if buy_is_crypto and not sell_is_crypto:
        if buy_amount is None:
            raise AppError("21bitcoin CSV has a BTC buy row with an empty buy_amount")
        direction = "inbound"
        asset = buy_currency
        amount = abs(buy_amount)
        fee = fee_amount if fee_currency == asset else Decimal("0")
        if sell_amount is not None and sell_currency:
            fiat_currency = sell_currency
            fiat_value = abs(sell_amount) + (fee_amount if fee_is_fiat else Decimal("0"))
        kind = "buy" if sell_currency else _twentyonebitcoin_kind(transaction_type, direction)
    elif sell_is_crypto and not buy_is_crypto:
        if sell_amount is None:
            raise AppError("21bitcoin CSV has a BTC sell/withdrawal row with an empty sell_amount")
        direction = "outbound"
        asset = sell_currency
        amount = abs(sell_amount)
        fee = fee_amount if fee_currency == asset else Decimal("0")
        if buy_amount is not None and buy_currency:
            fiat_currency = buy_currency
            fiat_value = max(Decimal("0"), abs(buy_amount) - (fee_amount if fee_is_fiat else Decimal("0")))
            kind = "sell"
        else:
            kind = _twentyonebitcoin_kind(transaction_type, direction)
    else:
        return None

    if fiat_value is not None and amount > 0:
        fiat_rate = fiat_value / amount
        pricing_source_kind = "exchange_execution"
        pricing_quality = "exact"

    note = str_or_none(_get_cell(sanitized, "note"))
    depot = str_or_none(_get_cell(sanitized, "depot_name"))
    description_parts = [part for part in (transaction_type, note, depot) if part]
    description = " - ".join(description_parts) or "Imported from 21bitcoin"
    provider_ref = f"21bitcoin:{row_id}"
    linked_transaction = str_or_none(_get_cell(sanitized, "linked_transaction"))
    external_ref = linked_transaction if is_l1_withdrawal and linked_transaction else provider_ref
    return {
        "txid": external_ref,
        "occurred_at": _twentyonebitcoin_datetime(_get_cell(sanitized, "transaction_date")),
        "direction": direction,
        "asset": asset,
        "amount": amount,
        "fee": fee,
        "fiat_rate": fiat_rate,
        "fiat_value": fiat_value,
        "fiat_currency": fiat_currency,
        "pricing_source_kind": pricing_source_kind,
        "pricing_provider": "21bitcoin",
        "pricing_pair": f"{asset}-{fiat_currency}" if fiat_currency else None,
        "pricing_method": "21bitcoin_csv" if pricing_source_kind else None,
        "pricing_external_ref": row_id,
        "pricing_quality": pricing_quality,
        "kind": kind,
        "description": description,
        "counterparty": str_or_none(_get_cell(sanitized, "exchange_name")) or "21bitcoin",
        "_exchange_evidence_matchable": is_l1_withdrawal,
        "_21bitcoin_l1_withdrawal": is_l1_withdrawal,
        "raw_json": json.dumps(json_ready(sanitized), sort_keys=True),
    }


def _twentyonebitcoin_kind(transaction_type, direction):
    tag = str(transaction_type or "").strip().lower()
    aliases = {
        "deposit": "deposit",
        "trade": "buy" if direction == "inbound" else "sell",
        "withdrawal": "withdrawal",
    }
    return aliases.get(tag, tag.replace(" ", "_") if tag else direction)


def load_twentyonebitcoin_csv_records(file_path):
    """Load 21bitcoin transaction-export CSV rows."""
    with open(file_path, "r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if not rows:
        return []
    header = {_normalized_column_key(column) for column in rows[0].keys()}
    missing = [
        column
        for column in _TWENTYONEBITCOIN_REQUIRED_COLUMNS
        if _normalized_column_key(column) not in header
    ]
    if missing:
        raise AppError("21bitcoin CSV is missing required columns: " + ", ".join(missing))
    normalized = []
    for row in rows:
        record = normalize_twentyonebitcoin_record(row)
        if record is not None:
            normalized.append(record)
    return normalized


def is_twentyonebitcoin_format(input_format):
    return input_format == "21bitcoin_csv"


# -- Strike ------------------------------------------------------------------


_STRIKE_REQUIRED_COLUMNS = (
    "Reference",
    "Date & Time (UTC)",
    "Transaction Type",
    "Amount BTC",
)


def _strike_datetime(value):
    text = str_or_none(value)
    if not text:
        return None
    text = text.strip()
    for fmt in ("%b %d %Y %H:%M:%S", "%B %d %Y %H:%M:%S"):
        try:
            dt = datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
        return dt.replace(microsecond=0).isoformat().replace("+00:00", "Z")
    return text


def _strike_fiat_currency(record):
    # Strike exports one fiat column family per CSV; pick that fiat from amount,
    # fee, or cost-basis headers and ignore timestamp suffixes such as "(UTC)".
    for key in record.keys():
        label = str(key or "").strip()
        label_key = label.casefold()
        if label_key in {"btc price", "amount btc", "fee btc"}:
            continue
        if "(" in label and ")" in label and label_key.startswith(("amount", "fee", "cost basis")):
            candidate = label.rsplit("(", 1)[1].split(")", 1)[0]
            currency = _currency_cell(candidate)
            if currency and currency not in {"BTC", "XBT", "LBTC"}:
                return currency
        parts = label.split()
        if len(parts) >= 2 and parts[0].casefold() in {"amount", "fee"}:
            currency = _currency_cell(parts[-1])
            if currency and currency not in {"BTC", "XBT", "LBTC"}:
                return currency
    return None


def _strike_lightning_row(transaction_type, destination):
    text = f"{transaction_type or ''} {destination or ''}".strip().casefold()
    destination_text = str(destination or "").strip().casefold()
    return (
        "lightning" in text
        or destination_text.startswith(("lnbc", "lntb", "lnbcrt", "lightning:"))
    )


def _strike_payment_hash(value, *, lightning):
    if not lightning:
        return None
    return _normalize_payment_hash(value)


def _strike_decimal_cell(value):
    text = str_or_none(value)
    if text and text.strip() in {"-", "–", "—"}:
        return None
    return _decimal_cell(value)


def _strike_kind(transaction_type, direction):
    tag = str(transaction_type or "").strip().casefold()
    aliases = {
        "buy": "buy",
        "purchase": "buy",
        "sell": "sell",
        "receive": "receive",
        "received": "receive",
        "send": "send",
        "sent": "send",
        "withdraw": "withdrawal",
        "withdrawal": "withdrawal",
    }
    return aliases.get(tag, tag.replace(" ", "_") if tag else direction)


def _strike_fiat_amount(record, fiat_currency):
    if not fiat_currency:
        return None
    return _strike_decimal_cell(
        _get_cell(
            record,
            f"Amount {fiat_currency}",
            f"Amount ({fiat_currency})",
        )
    )


def _strike_fiat_fee(record, fiat_currency):
    if not fiat_currency:
        return Decimal("0")
    return abs(
        _strike_decimal_cell(
            _get_cell(
                record,
                f"Fee {fiat_currency}",
                f"Fee ({fiat_currency})",
            )
        )
        or Decimal("0")
    )


def _strike_cost_basis(record, fiat_currency):
    if not fiat_currency:
        return None
    return _strike_decimal_cell(
        _get_cell(
            record,
            f"Cost Basis ({fiat_currency})",
            f"Cost Basis {fiat_currency}",
        )
    )


def normalize_strike_record(record):
    """Turn one Strike export row into the common import shape."""
    sanitized = {str(key).strip(): value for key, value in record.items() if key is not None}
    row_ref = str_or_none(_get_cell(sanitized, "Reference"))
    if not row_ref:
        raise AppError("Strike CSV has a BTC-side row without a Reference")

    amount_btc = _strike_decimal_cell(_get_cell(sanitized, "Amount BTC"))
    if amount_btc in (None, Decimal("0")):
        return None
    direction = "outbound" if amount_btc < 0 else "inbound"
    amount = abs(amount_btc)
    fee = abs(_strike_decimal_cell(_get_cell(sanitized, "Fee BTC")) or Decimal("0"))

    fiat_currency = _strike_fiat_currency(sanitized)
    amount_fiat = _strike_fiat_amount(sanitized, fiat_currency)
    fee_fiat = _strike_fiat_fee(sanitized, fiat_currency)
    cost_basis = _strike_cost_basis(sanitized, fiat_currency)
    fiat_rate = abs(_strike_decimal_cell(_get_cell(sanitized, "BTC Price")) or Decimal("0")) or None
    transaction_type = str_or_none(_get_cell(sanitized, "Transaction Type")) or "transaction"
    destination = str_or_none(_get_cell(sanitized, "Destination"))
    description = str_or_none(_get_cell(sanitized, "Description"))
    note = str_or_none(_get_cell(sanitized, "Note"))
    transaction_hash = str_or_none(_get_cell(sanitized, "Transaction Hash"))
    lightning = _strike_lightning_row(transaction_type, destination)
    payment_hash = _strike_payment_hash(transaction_hash, lightning=lightning)
    external_ref = f"strike:{row_ref}" if lightning or not transaction_hash else transaction_hash
    kind = _strike_kind(transaction_type, direction)

    fiat_value = None
    if amount_fiat is not None:
        if direction == "inbound" and kind == "buy":
            fiat_value = abs(amount_fiat) + fee_fiat
        elif direction == "outbound" and kind == "sell":
            fiat_value = max(Decimal("0"), abs(amount_fiat) - fee_fiat)
        else:
            fiat_value = abs(amount_fiat)
    elif fiat_rate is not None:
        fiat_value = amount * fiat_rate
    elif kind == "buy" and cost_basis is not None:
        fiat_value = abs(cost_basis)

    if fiat_rate is None and fiat_value is not None and amount > 0:
        fiat_rate = fiat_value / amount
    pricing_source_kind = "exchange_execution" if fiat_rate is not None or fiat_value is not None else None
    summary = " - ".join(part for part in (transaction_type, description, note) if part)
    raw_payload = dict(sanitized)
    raw_payload["_strike_lightning"] = lightning
    return {
        "txid": external_ref,
        "occurred_at": _strike_datetime(_get_cell(sanitized, "Date & Time (UTC)")),
        "direction": direction,
        "asset": "BTC",
        "amount": amount,
        "fee": fee,
        "fiat_rate": fiat_rate,
        "fiat_value": fiat_value,
        "fiat_currency": fiat_currency,
        "pricing_source_kind": pricing_source_kind,
        "pricing_provider": "Strike" if pricing_source_kind else None,
        "pricing_pair": f"BTC-{fiat_currency}" if fiat_currency and pricing_source_kind else None,
        "pricing_timestamp": _strike_datetime(_get_cell(sanitized, "Date & Time (UTC)"))
        if pricing_source_kind
        else None,
        "pricing_method": "strike_csv" if pricing_source_kind else None,
        "pricing_external_ref": row_ref if pricing_source_kind else None,
        "pricing_quality": "exact" if pricing_source_kind else None,
        "kind": kind,
        "description": summary or "Imported from Strike",
        "counterparty": "Strike",
        "payment_hash": payment_hash,
        "payment_hash_source": "importer" if payment_hash else None,
        "raw_json": json.dumps(json_ready(raw_payload), sort_keys=True),
    }


def load_strike_csv_records(file_path):
    """Load Strike transaction-history CSV rows."""
    with open(file_path, "r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if not rows:
        return []
    header = {_normalized_column_key(column) for column in rows[0].keys()}
    missing = [
        column
        for column in _STRIKE_REQUIRED_COLUMNS
        if _normalized_column_key(column) not in header
    ]
    if missing:
        raise AppError("Strike CSV is missing required columns: " + ", ".join(missing))
    normalized = []
    for row in rows:
        record = normalize_strike_record(row)
        if record is not None:
            normalized.append(record)
    return normalized


def is_strike_format(input_format):
    return input_format == "strike_csv"


# -- Ledger Live -------------------------------------------------------------


LEDGERLIVE_CSV_FORMAT = "ledgerlive_csv"
_LEDGERLIVE_REQUIRED_COLUMNS = (
    "Operation Date",
    "Currency Ticker",
    "Operation Type",
    "Operation Amount",
    "Operation Hash",
)
_LEDGERLIVE_REDACTED_COLUMNS = {
    "account xpub",
    "account xpub fingerprint",
    "xpub",
}


def _ledgerlive_sanitized_record(record):
    sanitized = {}
    for key, value in record.items():
        if key is None:
            continue
        column = str(key).strip()
        if _normalized_column_key(column) in _LEDGERLIVE_REDACTED_COLUMNS:
            sanitized[column] = "[redacted]"
        else:
            sanitized[column] = value
    return sanitized


def _ledgerlive_direction(value):
    tag = str(value or "").strip().casefold()
    if tag in {"in", "receive", "received"}:
        return "inbound", "deposit"
    if tag in {"out", "send", "sent"}:
        return "outbound", "withdrawal"
    return None, None


def normalize_ledgerlive_record(record, index=0):
    """Turn one Ledger Live operation-history row into wallet movement only."""
    sanitized = _ledgerlive_sanitized_record(record)
    asset = _bitcoin_family_asset(_get_cell(sanitized, "Currency Ticker"))
    if asset is None:
        return None
    direction, kind = _ledgerlive_direction(_get_cell(sanitized, "Operation Type"))
    if direction is None or kind is None:
        raise AppError(
            f"Ledger Live CSV has unsupported operation type '{_get_cell(sanitized, 'Operation Type')}'",
            code="validation",
            hint="Ledger Live imports support IN/OUT wallet movement rows for BTC/LBTC only.",
            retryable=False,
        )
    amount = _decimal_cell(_get_cell(sanitized, "Operation Amount"))
    if amount is None or amount == 0:
        fee = abs(_decimal_cell(_get_cell(sanitized, "Operation Fees")) or Decimal("0"))
        if fee > 0:
            return None
        raise AppError("Ledger Live CSV has a BTC row with an empty amount", code="validation")
    fee = abs(_decimal_cell(_get_cell(sanitized, "Operation Fees")) or Decimal("0"))
    occurred_at = _get_cell(sanitized, "Operation Date")
    txid = str_or_none(_get_cell(sanitized, "Operation Hash"))
    if not txid:
        txid = f"ledgerlive:{occurred_at}:{direction}:{asset}:{abs(amount)}:{index}"
    account_name = str_or_none(_get_cell(sanitized, "Account Name"))
    description = f"Ledger Live {kind}"
    if account_name:
        description = f"{description} - {account_name}"
    return {
        "txid": txid,
        "occurred_at": occurred_at,
        "direction": direction,
        "asset": asset,
        "amount": abs(amount),
        "fee": fee if direction == "outbound" else Decimal("0"),
        "kind": kind,
        "description": description,
        "counterparty": "Ledger Live",
        "raw_json": json.dumps(json_ready(sanitized), sort_keys=True),
    }


def load_ledgerlive_csv_records(file_path):
    """Load Ledger Live operation-history CSV rows.

    Ledger's fiat countervalues are intentionally ignored because the export
    labels them informational, not accounting-grade.
    """
    with open(file_path, "r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if not rows:
        return []
    header = {_normalized_column_key(column) for column in rows[0].keys()}
    missing = [
        column
        for column in _LEDGERLIVE_REQUIRED_COLUMNS
        if _normalized_column_key(column) not in header
    ]
    if missing:
        raise AppError("Ledger Live CSV is missing required columns: " + ", ".join(missing))
    normalized = []
    for index, row in enumerate(rows, start=1):
        record = normalize_ledgerlive_record(row, index=index)
        if record is not None:
            normalized.append(record)
    return normalized


def is_ledgerlive_format(input_format):
    return input_format == LEDGERLIVE_CSV_FORMAT


# -- Binance supplemental CSV ------------------------------------------------


BINANCE_SUPPLEMENTAL_CSV_FORMAT = "binance_supplemental_csv"
_BINANCE_SUPPLEMENTAL_AUTOINVEST_COLUMNS = (
    "timestamp utc",
    "base asset symbol",
    "quote asset amount + symbol",
    "trading fee (in quote asset)",
    "base asset amount + symbol",
)
_BINANCE_SUPPLEMENTAL_DIVIDEND_COLUMNS = (
    "id",
    "amount",
    "asset",
    "divtime",
)


def _split_amount_symbol(value, field):
    text = str_or_none(value)
    if text is None:
        return None, None
    parts = text.replace("\xa0", " ").strip().split()
    if len(parts) < 2:
        raise AppError(
            f"Binance supplemental CSV has invalid {field}",
            code="validation",
            hint=f"Expected '<amount> <asset>' in {field}.",
            retryable=False,
        )
    return _decimal_cell(parts[0]), _currency_cell(parts[1])


def _binance_supplemental_fee(value):
    text = str_or_none(value)
    if text is None or text.strip() in {"--", "-"}:
        return Decimal("0"), None
    return _split_amount_symbol(text, "trading fee")


def normalize_binance_supplemental_autoinvest_record(record, index=0):
    sanitized = {str(key).strip(): value for key, value in record.items() if key is not None}
    base_asset = _bitcoin_family_asset(
        _get_cell(sanitized, "base asset symbol", "Base Asset Symbol")
    )
    quote_amount, quote_asset = _split_amount_symbol(
        _get_cell(sanitized, "quote asset amount + symbol", "Quote Asset Amount + Symbol"),
        "quote asset amount",
    )
    base_amount, parsed_base_asset = _split_amount_symbol(
        _get_cell(sanitized, "base asset amount + symbol", "Base Asset Amount + Symbol"),
        "base asset amount",
    )
    fee_amount, fee_asset = _binance_supplemental_fee(
        _get_cell(sanitized, "trading fee (in quote asset)", "Trading Fee")
    )
    if base_asset is None and parsed_base_asset in _BITCOIN_FAMILY_ASSETS:
        base_asset = "BTC" if parsed_base_asset == "XBT" else parsed_base_asset
    if base_asset is None:
        return None
    if base_amount is None or quote_amount is None:
        raise AppError("Binance supplemental CSV has a BTC autoinvest row with an empty amount")
    if not _is_fiat_currency(quote_asset):
        raise AppError(
            "Binance supplemental BTC autoinvest rows must be funded by fiat for exact import",
            code="validation",
            hint=(
                "Crypto-funded Binance autoinvest rows are cross-asset trades; "
                "use the generic ledger with explicit review instead."
            ),
            retryable=False,
        )
    if fee_asset not in (None, quote_asset):
        raise AppError(
            "Binance supplemental BTC autoinvest fee asset does not match the quote asset",
            code="validation",
            retryable=False,
        )
    occurred_at = _get_cell(sanitized, "timestamp utc", "Timestamp UTC")
    source = str_or_none(_get_cell(sanitized, "source of funds", "Source of Funds"))
    fiat_value = abs(quote_amount) + abs(fee_amount or Decimal("0"))
    external_ref = (
        f"binance-supplemental:{occurred_at}:autoinvest:"
        f"{base_asset}:{base_amount}:{index}"
    )
    description = "Binance autoinvest buy"
    if source:
        description = f"{description} - {source}"
    return {
        "txid": external_ref,
        "occurred_at": occurred_at,
        "direction": "inbound",
        "asset": base_asset,
        "amount": abs(base_amount),
        "fee": Decimal("0"),
        "fiat_rate": fiat_value / abs(base_amount) if base_amount else None,
        "fiat_value": fiat_value,
        "fiat_currency": quote_asset,
        "pricing_source_kind": "exchange_execution",
        "pricing_provider": "Binance",
        "pricing_pair": f"{base_asset}-{quote_asset}",
        "pricing_timestamp": occurred_at,
        "pricing_method": "binance_supplemental_csv",
        "pricing_external_ref": external_ref,
        "pricing_quality": "exact",
        "kind": "buy",
        "description": description,
        "counterparty": "Binance",
        "raw_json": json.dumps(json_ready(sanitized), sort_keys=True),
    }


def _binance_ms_epoch_to_iso(value):
    text = str_or_none(value)
    if text is None:
        return None
    try:
        number = Decimal(text)
    except (ValueError, ArithmeticError) as exc:
        raise AppError("Binance supplemental CSV has invalid millisecond timestamp") from exc
    seconds = number / Decimal("1000")
    return (
        datetime.fromtimestamp(float(seconds), tz=timezone.utc)
        .replace(microsecond=0)
        .isoformat()
        .replace("+00:00", "Z")
    )


def normalize_binance_supplemental_dividend_record(record, index=0):
    sanitized = {str(key).strip(): value for key, value in record.items() if key is not None}
    asset = _bitcoin_family_asset(_get_cell(sanitized, "asset", "Asset"))
    if asset is None:
        return None
    amount = _decimal_cell(_get_cell(sanitized, "amount", "Amount"))
    if amount is None or amount == 0:
        return None
    ref = str_or_none(_get_cell(sanitized, "id", "tranId", "tran id")) or str(index)
    occurred_at = (
        _binance_ms_epoch_to_iso(_get_cell(sanitized, "divTime", "divtime", "time"))
        or _get_cell(sanitized, "timestamp", "Timestamp")
    )
    info = str_or_none(_get_cell(sanitized, "enInfo", "info", "description"))
    kind = "mining" if info and "mining" in info.casefold() else "income"
    return {
        "txid": f"binance:{ref}",
        "occurred_at": occurred_at,
        "direction": "inbound",
        "asset": asset,
        "amount": abs(amount),
        "fee": Decimal("0"),
        "kind": kind,
        "description": "Binance income" + (f" - {info}" if info else ""),
        "counterparty": "Binance",
        "raw_json": json.dumps(json_ready(sanitized), sort_keys=True),
    }


def load_binance_supplemental_csv_records(file_path):
    """Load BTC-relevant Binance supplemental CSV rows."""
    with open(file_path, "r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if not rows:
        return []
    header = {_normalized_column_key(column) for column in rows[0].keys()}
    autoinvest = all(
        _normalized_column_key(column) in header
        for column in _BINANCE_SUPPLEMENTAL_AUTOINVEST_COLUMNS
    )
    dividends = all(
        _normalized_column_key(column) in header
        for column in _BINANCE_SUPPLEMENTAL_DIVIDEND_COLUMNS
    )
    if not autoinvest and not dividends:
        raise AppError(
            "Binance supplemental CSV is not a supported BTC supplemental export",
            code="validation",
            hint=(
                "Use Binance autoinvest CSV or asset-dividend rows with id, "
                "amount, asset, divTime."
            ),
            retryable=False,
        )
    normalized = []
    for index, row in enumerate(rows, start=1):
        record = (
            normalize_binance_supplemental_autoinvest_record(row, index=index)
            if autoinvest
            else normalize_binance_supplemental_dividend_record(row, index=index)
        )
        if record is not None:
            normalized.append(record)
    return normalized


def is_binance_supplemental_format(input_format):
    return input_format == BINANCE_SUPPLEMENTAL_CSV_FORMAT


def is_exchange_evidence_format(input_format):
    return input_format in {
        "bullbitcoin_csv",
        "coinfinity_csv",
        "21bitcoin_csv",
        "pocketbitcoin_csv",
        "binance_supplemental_csv",
    }


def exchange_evidence_label(input_format):
    if input_format == "bullbitcoin_csv":
        return "Bull Bitcoin"
    if input_format == "coinfinity_csv":
        return "Coinfinity"
    if input_format == "21bitcoin_csv":
        return "21bitcoin"
    if input_format == "pocketbitcoin_csv":
        return "Pocket Bitcoin"
    if input_format == "binance_supplemental_csv":
        return "Binance"
    return str(input_format or "").replace("_", " ")


def exchange_evidence_rows_key(input_format):
    if input_format == "bullbitcoin_csv":
        return "bullbitcoin_rows"
    if input_format == "coinfinity_csv":
        return "coinfinity_rows"
    if input_format == "21bitcoin_csv":
        return "twentyonebitcoin_rows"
    if input_format == "pocketbitcoin_csv":
        return "pocketbitcoin_rows"
    if input_format == "binance_supplemental_csv":
        return "binance_rows"
    return "exchange_rows"


# -- River -------------------------------------------------------------------


def _river_fiat_amount(value, currency):
    if not currency or currency in {"BTC", "XBT"}:
        return None
    amount = _river_decimal(value)
    return abs(amount) if amount is not None else None


def _river_btc_amount(value, currency):
    if currency not in {"BTC", "XBT"}:
        return None
    amount = _river_decimal(value)
    return abs(amount) if amount is not None else None


def _river_kind(value):
    tag = str(value or "").strip().lower()
    aliases = {
        "automatic withdrawal": "withdrawal",
        "cash deposit": "cash_deposit",
        "cash withdrawal": "cash_withdrawal",
        "mining payout": "mining",
        "referral reward": "income",
        "bitcoin interest on cash": "interest",
    }
    return aliases.get(tag, tag.replace(" ", "_") if tag else "river_activity")


def normalize_river_record(record):
    """Turn one River CSV row into the common import-record shape.

    River currently documents two exports. Bitcoin Activity has the BTC/cash
    legs and a `Tag`; Account Activity adds reference code, transaction type,
    method/source/destination, price, and on-chain transaction id. Fiat-only
    rows are skipped by returning `None`; Kassiber is the BTC-side subledger.
    """
    sanitized = {str(key).strip(): value for key, value in record.items() if key is not None}
    sent_currency = _river_currency(_get_cell(sanitized, "Sent Currency"))
    received_currency = _river_currency(_get_cell(sanitized, "Received Currency"))
    fee_currency = _river_currency(_get_cell(sanitized, "Fee Currency"))
    price_currency = _river_currency(_get_cell(sanitized, "Bitcoin Price Currency"))
    sent_btc = _river_btc_amount(_get_cell(sanitized, "Sent Amount"), sent_currency)
    received_btc = _river_btc_amount(_get_cell(sanitized, "Received Amount"), received_currency)
    sent_fiat = _river_fiat_amount(_get_cell(sanitized, "Sent Amount"), sent_currency)
    received_fiat = _river_fiat_amount(_get_cell(sanitized, "Received Amount"), received_currency)
    fee_btc = _river_btc_amount(_get_cell(sanitized, "Fee Amount"), fee_currency) or Decimal("0")
    fee_fiat = _river_fiat_amount(_get_cell(sanitized, "Fee Amount"), fee_currency) or Decimal("0")
    price = _river_decimal(_get_cell(sanitized, "Bitcoin Price Amount"))
    reference = str_or_none(_get_cell(sanitized, "Transaction ID", "Reference Code"))
    tag = str_or_none(_get_cell(sanitized, "Tag", "Transaction Type"))
    transaction_type = str_or_none(_get_cell(sanitized, "Transaction Type")) or tag

    if received_btc is not None:
        direction = "inbound"
        amount = received_btc
        if sent_fiat is not None:
            fiat_value = sent_fiat + fee_fiat
            cash_leg_pricing = True
            kind = "buy"
        else:
            fiat_value = amount * price if price is not None else None
            cash_leg_pricing = False
            kind = _river_kind(transaction_type)
    elif sent_btc is not None:
        direction = "outbound"
        amount = sent_btc
        if received_fiat is not None:
            fiat_value = max(Decimal("0"), received_fiat - fee_fiat)
            cash_leg_pricing = True
            kind = "sell"
        else:
            fiat_value = amount * price if price is not None else None
            cash_leg_pricing = False
            kind = _river_kind(transaction_type)
    else:
        return None

    fiat_rate = fiat_value / amount if fiat_value is not None and amount > 0 else price
    fiat_currency = (
        sent_currency
        if sent_fiat is not None
        else received_currency
        if received_fiat is not None
        else price_currency
    )
    source = str_or_none(_get_cell(sanitized, "Source"))
    destination = str_or_none(_get_cell(sanitized, "Destination"))
    method = str_or_none(_get_cell(sanitized, "Method"))
    description_parts = [part for part in (transaction_type, method, source, destination) if part]
    description = " - ".join(description_parts) or "Imported from River"
    payment_hash = _normalize_payment_hash(_get_cell(sanitized, "Payment Hash", "payment_hash"))
    return {
        "txid": reference,
        "occurred_at": _get_cell(sanitized, "Date"),
        "direction": direction,
        "asset": "BTC",
        "amount": amount,
        "fee": fee_btc,
        "fiat_rate": fiat_rate,
        "fiat_value": fiat_value,
        "fiat_currency": fiat_currency,
        "pricing_source_kind": "exchange_execution"
        if cash_leg_pricing
        else "fmv_provider"
        if fiat_value is not None
        else None,
        "pricing_provider": "River",
        "pricing_pair": f"BTC-{fiat_currency}" if fiat_currency else None,
        "pricing_method": "river_csv",
        "pricing_external_ref": reference,
        "pricing_quality": "exact"
        if cash_leg_pricing
        else "provider_sample"
        if fiat_value is not None
        else None,
        "kind": kind,
        "description": description,
        "counterparty": destination if direction == "outbound" else source,
        "payment_hash": payment_hash,
        "payment_hash_source": "importer" if payment_hash else None,
        "_river_tag": tag,
        "_river_description": description,
        "raw_json": json.dumps(json_ready(sanitized), sort_keys=True),
    }


def load_river_csv_records(file_path):
    """Load River Bitcoin Activity or Account Activity CSV rows."""
    with open(file_path, "r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if not rows:
        return []
    header = {str(column).strip().casefold() for column in rows[0].keys()}
    missing = [column for column in _RIVER_REQUIRED_COLUMNS if column.casefold() not in header]
    if missing:
        raise AppError("River CSV is missing required columns: " + ", ".join(missing))
    normalized = []
    for row in rows:
        record = normalize_river_record(row)
        if record is not None:
            normalized.append(record)
    return normalized


def is_river_format(input_format):
    return input_format == "river_csv"


# -- Generic ledger (manual / generic tabular import) ------------------------
#
# A native, column-mapped import for a fill-in spreadsheet — the same shape an
# exchange "Excel import" template has, pared to Bitcoin scope. One Bitcoin leg
# per row (asset + amount + direction); the other side is the cash leg that
# becomes pricing. Reads .xlsx (via the optional `openpyxl` package) or
# CSV/TSV; `write_generic_ledger_template` emits the matching fill-in template.

GENERIC_LEDGER_FORMAT = "generic_ledger"

# Column headers for the fill-in template. Order defines the generated layout.
GENERIC_LEDGER_COLUMNS = (
    "Type",
    "Date",
    "Received Amount",
    "Received Asset",
    "Sent Amount",
    "Sent Asset",
    "Fee Amount",
    "Fee Asset",
    "Fiat Value",
    "Counterparty",
    "Note",
    "Tx-ID",
    "Payment Hash",
    "Payment Hash Source",
    "Swap Refund Funding Tx-ID",
)

# Asset codes (normalized via `normalize_asset_code`) treated as the Bitcoin
# leg. Anything else in an asset cell is treated as a fiat/cash currency (the
# pricing leg). `SATS`/`SAT` mark an amount denominated in whole satoshis.
_GENERIC_LEDGER_SATS_ASSETS = {"SATS", "SAT"}
_GENERIC_LEDGER_CRYPTO_ASSETS = {"BTC", "XBT", "LBTC"} | _GENERIC_LEDGER_SATS_ASSETS

# Type -> (direction, kind). `kind` is None for a plain acquisition/disposal.
# Every kind here is one the tax engine recognizes (see core/engines/rp2.py):
# income kinds map to RP2 earn transaction types; gift/donation/lost/stolen are
# deliberately routed to the non-sale-disposal quarantine for explicit review
# rather than being booked as ordinary market sales.
_GENERIC_LEDGER_TYPES = {
    "buy": ("inbound", "buy"),
    "purchase": ("inbound", "buy"),
    "sell": ("outbound", "sell"),
    "deposit": ("inbound", "deposit"),
    "receive": ("inbound", "deposit"),
    "transfer in": ("inbound", "deposit"),
    "withdrawal": ("outbound", "withdrawal"),
    "withdraw": ("outbound", "withdrawal"),
    "send": ("outbound", "withdrawal"),
    "transfer out": ("outbound", "withdrawal"),
    "spend": ("outbound", "spend"),
    "payment": ("outbound", "spend"),
    "income": ("inbound", "income"),
    "salary": ("inbound", "wages"),
    "wages": ("inbound", "wages"),
    "mining": ("inbound", "mining"),
    "staking": ("inbound", "staking"),
    "interest": ("inbound", "interest"),
    "lending interest": ("inbound", "lending_interest"),
    "airdrop": ("inbound", "airdrop"),
    "fork": ("inbound", "hardfork"),
    "hard fork": ("inbound", "hardfork"),
    "gift received": ("inbound", None),
    "gift sent": ("outbound", "gift"),
    "gift": ("outbound", "gift"),
    "donation": ("outbound", "donation"),
    "lost": ("outbound", "lost"),
    "stolen": ("outbound", "stolen"),
}

# Canonical Types shown in the template legend + dropdown, grouped for display.
GENERIC_LEDGER_TYPE_GROUPS = (
    ("Acquire", ("Buy", "Deposit", "Gift received")),
    ("Dispose", ("Sell", "Withdrawal", "Spend")),
    ("Earn", ("Income", "Mining", "Staking", "Interest", "Airdrop", "Fork")),
    ("Outflow (review)", ("Gift sent", "Donation", "Lost", "Stolen")),
)
GENERIC_LEDGER_DISPLAY_TYPES = tuple(
    label for _group, labels in GENERIC_LEDGER_TYPE_GROUPS for label in labels
)


def _generic_ledger_asset(value):
    code = str_or_none(value)
    return normalize_asset_code(code) if code else None


def _generic_ledger_is_crypto(asset):
    return bool(asset) and asset.upper() in _GENERIC_LEDGER_CRYPTO_ASSETS


def _generic_ledger_canonical_crypto(asset):
    upper = asset.upper() if asset else ""
    if upper in _GENERIC_LEDGER_SATS_ASSETS or upper == "XBT":
        return "BTC"
    if upper == "LBTC":
        return "LBTC"
    return asset


def _generic_ledger_btc_amount(amount, asset):
    """Coerce a leg amount to BTC Decimals, honoring a sats-denominated asset."""
    if amount is None:
        return None
    if asset and asset.upper() in _GENERIC_LEDGER_SATS_ASSETS:
        return abs(amount) / Decimal("100000000")
    return abs(amount)


# Header aliases that mark a row as the column header (and the columns the
# normalizer reads). Used so a sheet using "Transaction Type"/"Timestamp" is
# still recognized, matching what `normalize_generic_ledger_record` accepts.
_GENERIC_LEDGER_TYPE_HEADERS = {
    _normalized_column_key(name) for name in ("Type", "Transaction Type", "Kind")
}
_GENERIC_LEDGER_DATE_HEADERS = {
    _normalized_column_key(name) for name in ("Date", "Timestamp", "Time")
}

_GENERIC_LEDGER_CURRENCY_SYMBOLS = (
    "$", "€", "£", "CHF", "USD", "EUR", "GBP", "BTC", "XBT", "LBTC", "SATS", "SAT",
)
_GENERIC_LEDGER_EURO_DATE_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})(?:[ T](.+))?$")


def _generic_ledger_has_header(cells):
    seen = {_normalized_column_key(cell) for cell in cells if str_or_none(cell)}
    return bool(seen & _GENERIC_LEDGER_TYPE_HEADERS) and bool(
        seen & _GENERIC_LEDGER_DATE_HEADERS
    )


# Native template = Type + Date AND a Received/Sent leg column. A file with
# Type + Date but a single "Amount" column is NOT native — it takes the
# auto-detect (bring-your-own-file) path instead.
_GENERIC_LEDGER_NATIVE_AMOUNT_HEADERS = {
    _normalized_column_key(name)
    for name in ("Received Amount", "Buy Amount", "Sent Amount", "Sell Amount")
}


def _is_native_ledger(cells):
    seen = {_normalized_column_key(cell) for cell in cells if str_or_none(cell)}
    return _generic_ledger_has_header(cells) and bool(
        seen & _GENERIC_LEDGER_NATIVE_AMOUNT_HEADERS
    )


def _generic_ledger_decimal_cell(value, field, row_label):
    """Parse a hand-typed number, accepting both ``1,234.56`` and European
    ``1.234,56`` / ``0,05`` decimal-comma formats. Raises a row-numbered
    ``validation`` error rather than a context-free one on non-numeric input."""
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1].strip()
    for symbol in _GENERIC_LEDGER_CURRENCY_SYMBOLS:
        text = text.replace(symbol, "")
    text = text.replace("\xa0", "").replace(" ", "").strip()
    if not text:
        return None
    has_dot = "." in text
    has_comma = "," in text
    if has_dot and has_comma:
        # The right-most separator is the decimal point; the other groups thousands.
        if text.rfind(".") > text.rfind(","):
            text = text.replace(",", "")
        else:
            text = text.replace(".", "").replace(",", ".")
    elif has_comma:
        # A lone comma is a decimal separator; repeated commas group thousands.
        text = text.replace(",", "") if text.count(",") > 1 else text.replace(",", ".")
    elif text.count(".") > 1:
        # Repeated dots can only be thousands groupings (1.234.567).
        text = text.replace(".", "")
    if negative and not text.startswith("-"):
        text = "-" + text
    try:
        return dec(text)
    except AppError as exc:
        raise AppError(
            f"{row_label}: {field} is not a number ({value!r})",
            code="validation",
            hint="Use digits like 0.05000000 or 3000,00 — amounts are in BTC unless the asset is SATS.",
        ) from exc


def _generic_ledger_occurred_at(value, row_label):
    """Validate + normalize the Date cell. Accepts ISO/RFC3339 and Austrian
    ``DD.MM.YYYY`` (with optional time). Raises a row-numbered error otherwise,
    instead of letting a context-free error abort the whole import later."""
    text = str_or_none(value)
    if not text:
        raise AppError(
            f"{row_label}: missing Date",
            code="validation",
            hint="Use an ISO date such as 2026-01-15, a full timestamp, or DD.MM.YYYY.",
        )
    candidate = text.strip()
    match = _GENERIC_LEDGER_EURO_DATE_RE.match(candidate)
    if match:
        day, month, year, rest = match.groups()
        candidate = f"{year}-{int(month):02d}-{int(day):02d}"
        if rest:
            candidate = f"{candidate}T{rest.strip()}"
    try:
        parse_timestamp(candidate)
    except AppError as exc:
        raise AppError(
            f"{row_label}: Date '{text}' is not a recognized date",
            code="validation",
            hint="Use ISO 2026-01-15, a full RFC3339 timestamp, or DD.MM.YYYY.",
        ) from exc
    return candidate


def normalize_generic_ledger_record(record, index=0):
    """Turn one generic-ledger row into the common import record shape.

    Raises `AppError` with a row-numbered, actionable message on a malformed
    row rather than silently dropping it, since these rows are hand-entered.
    """
    sanitized = {str(key).strip(): value for key, value in record.items() if key is not None}
    row_label = f"Ledger row {index}" if index else "Ledger row"

    type_text = str_or_none(_get_cell(sanitized, "Type", "Transaction Type", "Kind"))
    if not type_text:
        raise AppError(
            f"{row_label}: missing Type",
            code="validation",
            hint="Set a Type such as Buy, Sell, Deposit, Withdrawal, Income, Mining, or Gift sent.",
        )
    type_key = " ".join(type_text.strip().lower().split())
    if type_key not in _GENERIC_LEDGER_TYPES:
        raise AppError(
            f"{row_label}: unknown Type '{type_text}'",
            code="validation",
            hint="Use one of: " + ", ".join(GENERIC_LEDGER_DISPLAY_TYPES) + ".",
        )
    type_direction, kind = _GENERIC_LEDGER_TYPES[type_key]

    received_asset = _generic_ledger_asset(
        _get_cell(sanitized, "Received Asset", "Received Cur.", "Buy Asset", "Buy Cur.")
    )
    sent_asset = _generic_ledger_asset(
        _get_cell(sanitized, "Sent Asset", "Sent Cur.", "Sell Asset", "Sell Cur.")
    )
    received_amount = _generic_ledger_decimal_cell(
        _get_cell(sanitized, "Received Amount", "Buy Amount"), "Received Amount", row_label
    )
    sent_amount = _generic_ledger_decimal_cell(
        _get_cell(sanitized, "Sent Amount", "Sell Amount"), "Sent Amount", row_label
    )
    fee_amount = _generic_ledger_decimal_cell(
        _get_cell(sanitized, "Fee Amount", "Fee"), "Fee Amount", row_label
    )
    fee_asset = _generic_ledger_asset(_get_cell(sanitized, "Fee Asset", "Fee Cur."))
    explicit_fiat = _generic_ledger_decimal_cell(
        _get_cell(
            sanitized,
            "Fiat Value",
            "Value (fiat)",
            "Buy Value (fiat)",
            "Sell Value (fiat)",
        ),
        "Fiat Value",
        row_label,
    )

    received_is_crypto = _generic_ledger_is_crypto(received_asset) and received_amount is not None
    sent_is_crypto = _generic_ledger_is_crypto(sent_asset) and sent_amount is not None
    if received_is_crypto and sent_is_crypto:
        raise AppError(
            f"{row_label}: both sides are Bitcoin — crypto-to-crypto rows are not supported",
            code="validation",
            hint="Record one Bitcoin leg per row and price it with a fiat amount or a Fiat Value.",
        )
    if not received_is_crypto and not sent_is_crypto:
        raise AppError(
            f"{row_label}: no Bitcoin leg found (expected BTC, LBTC, or SATS on the Received or Sent side)",
            code="validation",
            hint="Put the Bitcoin amount + asset on the Received side (inbound) or the Sent side (outbound).",
        )

    if received_is_crypto:
        direction = "inbound"
        leg_asset = received_asset
        amount = _generic_ledger_btc_amount(received_amount, received_asset)
        fiat_asset = sent_asset if (sent_asset and not _generic_ledger_is_crypto(sent_asset)) else None
        fiat_leg_amount = sent_amount if fiat_asset else None
    else:
        direction = "outbound"
        leg_asset = sent_asset
        amount = _generic_ledger_btc_amount(sent_amount, sent_asset)
        fiat_asset = received_asset if (received_asset and not _generic_ledger_is_crypto(received_asset)) else None
        fiat_leg_amount = received_amount if fiat_asset else None
    # `leg_asset` keeps the row's original Bitcoin spelling (e.g. SATS); `asset`
    # is canonicalized for storage (SATS/XBT -> BTC). The fee fallback below
    # needs the original so a blank Fee Asset on a SATS leg stays in sats.
    asset = _generic_ledger_canonical_crypto(leg_asset)

    if type_direction != direction:
        raise AppError(
            f"{row_label}: Type '{type_text}' is {type_direction} but the Bitcoin leg is {direction}",
            code="validation",
            hint=(
                "Inbound Types (Buy/Deposit/Income/…) need the Bitcoin amount on the Received side; "
                "outbound Types (Sell/Withdrawal/Spend/Gift sent/…) on the Sent side."
            ),
        )

    fee = Decimal("0")
    fee_fiat = Decimal("0")
    if fee_amount is not None and abs(fee_amount) > 0:
        if _generic_ledger_is_crypto(fee_asset) or fee_asset is None:
            # No fee asset given defaults to the Bitcoin leg (on-chain/network
            # fee), using the leg's ORIGINAL spelling so a blank fee on a SATS
            # leg is read as sats, not BTC.
            fee = _generic_ledger_btc_amount(abs(fee_amount), fee_asset or leg_asset)
        else:
            # A fiat fee only makes sense as an adjustment to a same-currency
            # cash leg's execution price; network fees belong on the Bitcoin leg.
            if not (fiat_asset and fiat_leg_amount is not None):
                raise AppError(
                    f"{row_label}: a fiat fee ({fee_asset}) needs a matching cash leg",
                    code="validation",
                    hint=(
                        "Put network fees in BTC (Fee Asset BTC or blank), or add the "
                        "Sent/Received fiat amount this fee applies to."
                    ),
                )
            if fee_asset != fiat_asset:
                raise AppError(
                    f"{row_label}: fee currency {fee_asset} does not match the trade currency {fiat_asset}",
                    code="validation",
                    hint="Use the same currency for the fee and the row's fiat amount.",
                )
            fee_fiat = abs(fee_amount)

    fiat_currency = None
    fiat_value = None
    pricing_source_kind = None
    pricing_quality = None
    if fiat_leg_amount is not None and fiat_asset:
        fiat_currency = fiat_asset
        base = abs(fiat_leg_amount)
        if direction == "inbound":
            fiat_value = base + fee_fiat
        else:
            fiat_value = max(Decimal("0"), base - fee_fiat)
        pricing_source_kind = "exchange_execution"
        pricing_quality = "exact"
    elif explicit_fiat is not None:
        # Declared fair-market value in the book currency (no cash leg). Currency
        # is left unset so it is treated as the book currency downstream.
        fiat_value = abs(explicit_fiat)

    fiat_rate = None
    if fiat_value is not None and amount and amount > 0:
        fiat_rate = fiat_value / amount

    occurred_at = _generic_ledger_occurred_at(
        _get_cell(sanitized, "Date", "Timestamp", "Time"), row_label
    )

    # No synthetic positional id: an absent Tx-ID leaves external_id empty so
    # dedup falls back to the economic fingerprint (occurred_at/direction/asset/
    # amount/fee). A row-index-based id would shift on insert/reorder and
    # re-import the same rows as duplicates.
    txid = str_or_none(_get_cell(sanitized, "Tx-ID", "TxID", "Txid", "Transaction ID"))
    external_id = txid or ""
    note = str_or_none(_get_cell(sanitized, "Note", "Comment", "Description"))
    counterparty = str_or_none(_get_cell(sanitized, "Counterparty", "Exchange", "Platform"))
    payment_hash = str_or_none(
        _get_cell(
            sanitized,
            "Payment Hash",
            "PaymentHash",
            "Lightning Payment Hash",
            "LN Payment Hash",
        )
    )
    payment_hash_source = str_or_none(
        _get_cell(
            sanitized,
            "Payment Hash Source",
            "PaymentHash Source",
            "Payment Hash Origin",
        )
    )
    swap_refund_funding_txid = str_or_none(
        _get_cell(
            sanitized,
            "Swap Refund Funding Tx-ID",
            "Swap Refund Funding TxID",
            "Refund Funding Tx-ID",
            "Refund Funding TxID",
            "HTLC Funding TxID",
        )
    )

    return {
        "txid": external_id,
        "occurred_at": occurred_at,
        "direction": direction,
        "asset": asset,
        "amount": amount,
        "fee": fee,
        "fiat_rate": fiat_rate,
        "fiat_value": fiat_value,
        "fiat_currency": fiat_currency,
        "pricing_source_kind": pricing_source_kind,
        "pricing_provider": counterparty if pricing_source_kind else None,
        "pricing_pair": f"{asset}-{fiat_currency}" if (pricing_source_kind and fiat_currency) else None,
        "pricing_method": "generic_ledger" if pricing_source_kind else None,
        "pricing_external_ref": txid or None,
        "pricing_quality": pricing_quality,
        "kind": kind,
        "description": note,
        "counterparty": counterparty,
        "payment_hash": payment_hash,
        "payment_hash_source": payment_hash_source or ("generic_ledger" if payment_hash else None),
        "swap_refund_funding_txid": swap_refund_funding_txid,
        "raw_json": json.dumps(json_ready(sanitized), sort_keys=True),
    }


def _generic_ledger_xlsx_cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _pick_generic_ledger_sheet(workbook):
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
            if _generic_ledger_has_header(row):
                return sheet
    return workbook.active


def _read_generic_ledger_xlsx(file_path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised only without openpyxl
        raise AppError(
            "Reading .xlsx ledger files needs the optional 'openpyxl' package",
            code="dependency_missing",
            hint="Install kassiber with openpyxl, or export the ledger as CSV and import that instead.",
        ) from exc
    try:
        workbook = load_workbook(file_path, data_only=True)
    except Exception as exc:  # openpyxl raises assorted errors on malformed files
        raise AppError(
            f"Could not read .xlsx ledger file: {exc}",
            code="validation",
            hint="Re-export the file as .xlsx or CSV and try again.",
        ) from exc
    try:
        sheet = _pick_generic_ledger_sheet(workbook)
        rows = [
            [_generic_ledger_xlsx_cell(cell) for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
    finally:
        workbook.close()
    return rows


def _read_generic_ledger_csv(file_path):
    with open(file_path, "r", encoding="utf-8-sig", newline="") as handle:
        text = handle.read()
    if not text.strip():
        return []
    try:
        dialect: Any = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
    except csv.Error:
        dialect = None
    reader = (
        csv.reader(io.StringIO(text), dialect)
        if dialect is not None
        else csv.reader(io.StringIO(text))
    )
    return [list(row) for row in reader]


def _generic_ledger_rows_to_records(rows):
    header_index = None
    for index, row in enumerate(rows):
        if _generic_ledger_has_header(row):
            header_index = index
            break
    if header_index is None:
        raise AppError(
            "Ledger file has no recognizable header row (expected at least 'Type' and 'Date' columns)",
            code="validation",
            hint="Download the import template with `wallets ledger-template` and keep its header row.",
        )
    header = [str(cell).strip() if cell is not None else "" for cell in rows[header_index]]
    records = []
    for row in rows[header_index + 1 :]:
        if not any(str_or_none(cell) for cell in row):
            continue
        record = {}
        for column_index, name in enumerate(header):
            if not name:
                continue
            record[name] = row[column_index] if column_index < len(row) else ""
        records.append(record)
    return records


# --------------------------------------------------------------------------- #
# Bring-your-own-file: auto-detect arbitrary column layouts onto the ledger
# shape so a non-template export still imports through the same normalizer (and
# its Type -> (direction, kind) taxonomy + exact fiat pricing) untouched. Only
# files WITHOUT a native Type+Date header take this path; template files are
# unchanged.
# --------------------------------------------------------------------------- #
_BYO_DATE = {_normalized_column_key(n) for n in ("Date", "Time", "Timestamp", "Datetime", "Date Time", "Datum", "Zeitpunkt", "Executed At", "Trade Date")}
_BYO_TYPE = {_normalized_column_key(n) for n in ("Type", "Transaction Type", "Tx Type", "Side", "Category", "Action", "Art")}
_BYO_DIRECTION = {_normalized_column_key(n) for n in ("Direction", "In/Out", "Flow", "Richtung")}
_BYO_RECEIVED = {_normalized_column_key(n) for n in ("Received Amount", "Received", "Received BTC", "Buy Amount", "Incoming", "Amount Received", "Deposit Amount", "Credit", "Eingang", "Erhalten")}
_BYO_SENT = {_normalized_column_key(n) for n in ("Sent Amount", "Sent", "Sent BTC", "Sell Amount", "Outgoing", "Amount Sent", "Withdrawal Amount", "Debit", "Ausgang", "Gesendet")}
_BYO_AMOUNT = {_normalized_column_key(n) for n in ("Amount", "BTC", "Amount BTC", "BTC Amount", "Quantity", "Qty", "Net Amount", "Betrag", "Menge")}
_BYO_RECEIVED_ASSET = {_normalized_column_key(n) for n in ("Received Asset", "Received Currency", "Buy Asset", "Buy Currency", "Incoming Asset", "Credit Asset")}
_BYO_SENT_ASSET = {_normalized_column_key(n) for n in ("Sent Asset", "Sent Currency", "Sell Asset", "Sell Currency", "Outgoing Asset", "Debit Asset")}
_BYO_FEE_ASSET = {_normalized_column_key(n) for n in ("Fee Asset", "Fee Currency", "Fee Cur.", "Fee Coin")}
# "Currency" is fiat far more often than a crypto-asset column, so it belongs
# to fiat detection; the crypto asset is only taken from explicit asset columns.
_BYO_ASSET = {_normalized_column_key(n) for n in ("Asset", "Coin", "Symbol", "Crypto", "Crypto Asset")}
_BYO_FEE = {_normalized_column_key(n) for n in ("Fee", "Fees", "Fee Amount", "Miner Fee", "Network Fee", "Tx Fee", "Gebühr", "Gebuehr")}
_BYO_FIAT_CURRENCY = {_normalized_column_key(n) for n in ("Fiat Currency", "Fiat", "Currency", "Quote Currency", "Cash Currency", "Währung", "Waehrung")}
_BYO_FIAT_VALUE = {_normalized_column_key(n) for n in ("Fiat Value", "Total", "Total Value", "Value", "Proceeds", "Cost", "Gesamt", "Wert")}
_BYO_FIAT_RATE = {_normalized_column_key(n) for n in ("Price", "Rate", "Unit Price", "Price Per BTC", "BTC Price", "Spot", "Kurs", "Preis")}
_BYO_NOTE = {_normalized_column_key(n) for n in ("Note", "Notes", "Description", "Memo", "Label", "Comment", "Notiz", "Beschreibung")}
_BYO_TXID = {_normalized_column_key(n) for n in ("Tx-ID", "TxID", "Txid", "Transaction ID", "Tx Hash", "Hash", "Reference", "Ref")}
_BYO_PAYMENT_HASH = {_normalized_column_key(n) for n in ("Payment Hash", "PaymentHash", "Lightning Payment Hash", "LN Payment Hash")}
_BYO_PAYMENT_HASH_SOURCE = {_normalized_column_key(n) for n in ("Payment Hash Source", "PaymentHash Source", "Payment Hash Origin")}
_BYO_SWAP_REFUND_FUNDING_TXID = {
    _normalized_column_key(n)
    for n in (
        "Swap Refund Funding Tx-ID",
        "Swap Refund Funding TxID",
        "Refund Funding Tx-ID",
        "Refund Funding TxID",
        "HTLC Funding TxID",
    )
}
_BYO_COUNTERPARTY = {_normalized_column_key(n) for n in ("Counterparty", "Exchange", "Platform", "Payee", "Gegenpartei")}

_BYO_TYPE_VALUE_MAP = {
    "in": "Deposit", "incoming": "Deposit", "received": "Deposit", "receive": "Deposit",
    "credit": "Deposit", "out": "Withdrawal", "outgoing": "Withdrawal", "sent": "Withdrawal",
    "send": "Withdrawal", "debit": "Withdrawal",
}
_BYO_INBOUND_VALUES = {"in", "inbound", "received", "receive", "incoming", "credit", "deposit", "buy"}
_BYO_OUTBOUND_VALUES = {"out", "outbound", "sent", "send", "outgoing", "debit", "withdrawal", "sell"}
_BYO_RECEIVED_HEADER_TOKENS = {"received", "receive", "incoming", "credit", "deposit", "buy", "bought", "erhalten", "eingang"}
_BYO_SENT_HEADER_TOKENS = {"sent", "send", "outgoing", "debit", "withdrawal", "withdraw", "sell", "sold", "ausgang", "gesendet"}
_BYO_HEADER_ASSET_ALIASES = {
    "btc": "BTC",
    "bitcoin": "BTC",
    "xbt": "BTC",
    "lbtc": "LBTC",
    "liquidbtc": "LBTC",
    "liquid": "LBTC",
    "sats": "SATS",
    "sat": "SATS",
    "eur": "EUR",
    "euro": "EUR",
    "usd": "USD",
    "dollar": "USD",
    "chf": "CHF",
    "gbp": "GBP",
    "aud": "AUD",
    "cad": "CAD",
    "jpy": "JPY",
    "nok": "NOK",
    "sek": "SEK",
    "dkk": "DKK",
    "pln": "PLN",
    "czk": "CZK",
    "huf": "HUF",
}


def _byo_number(cell):
    """Lenient numeric parse (honors locale decimals); None on blank/junk."""
    try:
        return _generic_ledger_decimal_cell(cell, "x", "")
    except AppError:
        return None


def _byo_header_tokens(value):
    return [token for token in re.split(r"[^a-z0-9]+", str(value or "").casefold()) if token]


def _byo_asset_from_header(value):
    tokens = _byo_header_tokens(value)
    if not tokens:
        return None
    joined = " ".join(tokens)
    if "liquid bitcoin" in joined:
        return "LBTC"
    for token in reversed(tokens):
        mapped = _BYO_HEADER_ASSET_ALIASES.get(token)
        if mapped:
            return mapped
    return None


def _byo_leg_role_from_header(value):
    tokens = set(_byo_header_tokens(value))
    if tokens & _BYO_RECEIVED_HEADER_TOKENS:
        return "received"
    if tokens & _BYO_SENT_HEADER_TOKENS:
        return "sent"
    return None


def _byo_type_value(type_value, direction, *, has_cash_counterleg=False):
    if type_value:
        return _BYO_TYPE_VALUE_MAP.get(_normalized_column_key(type_value), type_value)
    if has_cash_counterleg:
        return "Buy" if direction == "inbound" else "Sell"
    return "Deposit" if direction == "inbound" else "Withdrawal"


def _byo_type_direction(type_value):
    mapped = _byo_type_value(type_value, None)
    entry = _GENERIC_LEDGER_TYPES.get(" ".join(str(mapped).strip().lower().split()))
    return entry[0] if entry else None


def _byo_direction_value(value):
    key = _normalized_column_key(value or "")
    if key in _BYO_INBOUND_VALUES:
        return "inbound"
    if key in _BYO_OUTBOUND_VALUES:
        return "outbound"
    return None


def infer_ledger_columns(header):
    """Guess a column plan mapping an arbitrary header onto the ledger shape.

    Returns ``{"plan", "detected", "confident"}``. ``confident`` requires a date
    column and a usable amount layout (received/sent, or a single amount with a
    type/direction column, or a signed amount). Used only for non-template files.
    """
    norm = {h: _normalized_column_key(h) for h in header if str_or_none(h)}
    used = set()
    detected = []

    def take(aliases, field):
        for original in header:
            if original in used or original not in norm:
                continue
            if norm[original] in aliases:
                used.add(original)
                detected.append({"column": original, "field": field})
                return original
        return None

    def take_asset_suffixed_leg(field):
        for original in header:
            if original in used or original not in norm:
                continue
            if _byo_leg_role_from_header(original) == field and _byo_asset_from_header(original):
                used.add(original)
                detected.append({"column": original, "field": field})
                return original
        return None

    date = take(_BYO_DATE, "date")
    received = take(_BYO_RECEIVED, "received") or take_asset_suffixed_leg("received")
    sent = take(_BYO_SENT, "sent") or take_asset_suffixed_leg("sent")
    amount = None if (received and sent) else take(_BYO_AMOUNT, "amount")
    type_col = take(_BYO_TYPE, "type")
    direction = take(_BYO_DIRECTION, "direction") if not type_col else None
    received_asset = take(_BYO_RECEIVED_ASSET, "received_asset")
    sent_asset = take(_BYO_SENT_ASSET, "sent_asset")
    asset = take(_BYO_ASSET, "asset")
    fee = take(_BYO_FEE, "fee")
    fee_asset = take(_BYO_FEE_ASSET, "fee_asset")
    fiat_currency = take(_BYO_FIAT_CURRENCY, "fiat_currency")
    fiat_value = take(_BYO_FIAT_VALUE, "fiat_value")
    fiat_rate = take(_BYO_FIAT_RATE, "fiat_rate")
    note = take(_BYO_NOTE, "description")
    txid = take(_BYO_TXID, "txid")
    payment_hash = take(_BYO_PAYMENT_HASH, "payment_hash")
    payment_hash_source = take(_BYO_PAYMENT_HASH_SOURCE, "payment_hash_source")
    swap_refund_funding_txid = take(_BYO_SWAP_REFUND_FUNDING_TXID, "swap_refund_funding_txid")
    counterparty = take(_BYO_COUNTERPARTY, "counterparty")

    plan = {
        "date": date, "type": type_col, "direction": direction,
        "received": received, "sent": sent, "amount": amount,
        "received_asset": received_asset, "sent_asset": sent_asset,
        "asset": asset, "fee": fee, "fee_asset": fee_asset,
        "fiat_currency": fiat_currency, "fiat_value": fiat_value,
        "fiat_rate": fiat_rate, "note": note, "txid": txid, "counterparty": counterparty,
        "payment_hash": payment_hash, "payment_hash_source": payment_hash_source,
        "swap_refund_funding_txid": swap_refund_funding_txid,
        "received_header_asset": _byo_asset_from_header(received),
        "sent_header_asset": _byo_asset_from_header(sent),
        "amount_header_asset": _byo_asset_from_header(amount),
        "fee_header_asset": _byo_asset_from_header(fee),
    }
    confident = bool(date) and bool(received or sent or amount)
    return {"plan": plan, "detected": detected, "confident": confident}


def _ledger_plan_usable(plan):
    return bool(plan) and bool(plan.get("date")) and bool(
        plan.get("received") or plan.get("sent") or plan.get("amount")
    )


def _remap_byo_row_to_ledger(row, plan):
    """Remap one arbitrary row into a #244-shaped ledger record (string cells).

    Only routes raw cell values into the right ledger columns; all parsing and
    validation stays in ``normalize_generic_ledger_record``.
    """
    def cell(col):
        return str_or_none(row.get(col)) if col else None

    def crypto(asset_code):
        return _generic_ledger_is_crypto(_generic_ledger_asset(asset_code))

    out = {"Date": cell(plan.get("date")) or ""}
    asset = cell(plan.get("asset")) or plan.get("amount_header_asset") or "BTC"
    received_asset = (
        cell(plan.get("received_asset"))
        or plan.get("received_header_asset")
        or asset
    )
    sent_asset = (
        cell(plan.get("sent_asset"))
        or plan.get("sent_header_asset")
        or asset
    )
    type_value = cell(plan.get("type"))

    direction = None
    btc_cell = None
    has_cash_counterleg = False
    both_amounts_present = False
    if plan.get("received") or plan.get("sent"):
        received_num = _byo_number(row.get(plan.get("received"))) if plan.get("received") else None
        sent_num = _byo_number(row.get(plan.get("sent"))) if plan.get("sent") else None
        received_present = received_num is not None and received_num != 0
        sent_present = sent_num is not None and sent_num != 0
        if received_present and not sent_present:
            direction, btc_cell = "inbound", cell(plan["received"])
        elif sent_present and not received_present:
            direction, btc_cell = "outbound", cell(plan["sent"])
        elif received_present and sent_present:
            both_amounts_present = True
            received_is_crypto = crypto(received_asset)
            sent_is_crypto = crypto(sent_asset)
            if received_is_crypto and not sent_is_crypto:
                direction, btc_cell = "inbound", cell(plan["received"])
                has_cash_counterleg = True
            elif sent_is_crypto and not received_is_crypto:
                direction, btc_cell = "outbound", cell(plan["sent"])
                has_cash_counterleg = True
    elif plan.get("amount"):
        raw = cell(plan["amount"])
        number = _byo_number(row.get(plan["amount"]))
        if type_value:
            direction = _byo_type_direction(type_value)
            btc_cell = format(abs(number), "f") if (number is not None and direction) else raw
        elif plan.get("direction"):
            direction = _byo_direction_value(cell(plan["direction"]))
            btc_cell = format(abs(number), "f") if number is not None else raw
        elif number is not None:
            direction = "outbound" if number < 0 else "inbound"
            btc_cell = format(abs(number), "f")
        else:
            direction, btc_cell = "inbound", raw

    if both_amounts_present:
        out["Type"] = _byo_type_value(
            type_value,
            direction,
            has_cash_counterleg=has_cash_counterleg,
        )
        if plan.get("received"):
            out["Received Asset"], out["Received Amount"] = received_asset, cell(plan["received"])
        if plan.get("sent"):
            out["Sent Asset"], out["Sent Amount"] = sent_asset, cell(plan["sent"])
        _byo_passthrough(out, row, plan)
        fee_cell = cell(plan.get("fee"))
        if fee_cell is not None:
            out["Fee Amount"] = fee_cell
            fee_asset = cell(plan.get("fee_asset")) or plan.get("fee_header_asset")
            if fee_asset:
                out["Fee Asset"] = fee_asset
        return out

    if plan.get("received") and direction == "inbound":
        asset = received_asset
    elif plan.get("sent") and direction == "outbound":
        asset = sent_asset

    out["Type"] = _byo_type_value(type_value, direction)

    fiat_currency = cell(plan.get("fiat_currency"))
    fiat_value = cell(plan.get("fiat_value"))
    if fiat_value is None and plan.get("fiat_rate") and btc_cell is not None:
        rate = _byo_number(row.get(plan["fiat_rate"]))
        magnitude = _byo_number(btc_cell)
        if rate is not None and magnitude is not None:
            fiat_value = format(rate * magnitude, "f")

    if direction == "outbound":
        out["Sent Asset"], out["Sent Amount"] = asset, btc_cell
        if fiat_value is not None:
            if fiat_currency:
                out["Received Asset"], out["Received Amount"] = fiat_currency, fiat_value
            else:
                out["Fiat Value"] = fiat_value
    else:
        out["Received Asset"], out["Received Amount"] = asset, btc_cell
        if fiat_value is not None:
            if fiat_currency:
                out["Sent Asset"], out["Sent Amount"] = fiat_currency, fiat_value
            else:
                out["Fiat Value"] = fiat_value

    fee_cell = cell(plan.get("fee"))
    if fee_cell is not None:
        out["Fee Amount"] = fee_cell
        fee_asset = cell(plan.get("fee_asset")) or plan.get("fee_header_asset")
        if fee_asset:
            out["Fee Asset"] = fee_asset
    _byo_passthrough(out, row, plan)
    return out


def _byo_passthrough(out, row, plan):
    if plan.get("txid"):
        out["Tx-ID"] = str_or_none(row.get(plan["txid"]))
    if plan.get("payment_hash"):
        out["Payment Hash"] = str_or_none(row.get(plan["payment_hash"]))
    if plan.get("payment_hash_source"):
        out["Payment Hash Source"] = str_or_none(row.get(plan["payment_hash_source"]))
    if plan.get("swap_refund_funding_txid"):
        out["Swap Refund Funding Tx-ID"] = str_or_none(row.get(plan["swap_refund_funding_txid"]))
    if plan.get("note"):
        out["Note"] = str_or_none(row.get(plan["note"]))
    if plan.get("counterparty"):
        out["Counterparty"] = str_or_none(row.get(plan["counterparty"]))


def _read_ledger_rows(file_path):
    if not os.path.exists(file_path):
        raise AppError(
            f"Import file not found: {file_path}",
            code="not_found",
            hint="Check the file path.",
        )
    extension = os.path.splitext(file_path)[1].lower()
    if extension in {".xlsx", ".xlsm"}:
        return _read_generic_ledger_xlsx(file_path)
    return _read_generic_ledger_csv(file_path)


def _ledger_source_records(rows, column_map=None):
    """Return #244-shaped record dicts from raw rows.

    Native template files (Type+Date header) use the existing path unchanged.
    Other files are auto-detected (or use an explicit ``column_map`` plan) and
    remapped onto the ledger shape, raising ``ledger_unrecognized`` when the
    columns can't be recognized.
    """
    header_row = next((row for row in rows if any(str_or_none(cell) for cell in row)), None)
    if column_map is None and header_row is not None and _is_native_ledger(header_row):
        return _generic_ledger_rows_to_records(rows), None

    header = [str(cell).strip() if cell is not None else "" for cell in (header_row or [])]
    if column_map is not None:
        plan, detected = column_map, None
    else:
        inferred = infer_ledger_columns(header)
        plan, detected = inferred["plan"], inferred["detected"]
    if not _ledger_plan_usable(plan):
        raise AppError(
            "Could not recognize the columns in this file.",
            code="ledger_unrecognized",
            hint="Download the import template (it already has the right columns), fill it in, and import that — or map the columns yourself.",
            details={"headers": header},
        )
    header_index = rows.index(header_row)
    records = []
    for raw in rows[header_index + 1 :]:
        if not any(str_or_none(cell) for cell in raw):
            continue
        row = {name: (raw[i] if i < len(raw) else "") for i, name in enumerate(header) if name}
        records.append(_remap_byo_row_to_ledger(row, plan))
    return records, detected


def load_generic_ledger_records(file_path, column_map=None):
    """Load a generic-ledger .xlsx or CSV/TSV file into common import records.

    Template files import as before; arbitrary files are auto-detected and
    remapped onto the ledger shape (``column_map`` overrides the guess).
    """
    rows = _read_ledger_rows(file_path)
    records, _ = _ledger_source_records(rows, column_map)
    normalized = [
        normalize_generic_ledger_record(record, index=index)
        for index, record in enumerate(records, start=1)
    ]
    if not normalized:
        raise AppError(
            "Ledger file has a header but no transaction rows to import",
            code="validation",
            hint="Add at least one transaction row below the header.",
        )
    return normalized


def _generic_ledger_preview_row(record):
    """A JSON-safe subset of a normalized ledger record for preview display."""
    def _safe(value):
        return format(value, "f") if isinstance(value, Decimal) else value

    return {
        "occurred_at": record.get("occurred_at"),
        "direction": record.get("direction"),
        "asset": record.get("asset"),
        "amount": _safe(record.get("amount")),
        "fee": _safe(record.get("fee")),
        "kind": record.get("kind"),
        "fiat_currency": record.get("fiat_currency"),
        "fiat_value": _safe(record.get("fiat_value")),
        "description": record.get("description"),
    }


def preview_generic_ledger_records(file_path, *, limit=200, column_map=None):
    """Report what a generic-ledger file would import, without persisting.

    Reuses the same reader + per-row normalizer as the real import, but catches
    each row's validation error individually so the whole file previews at once
    (the importer itself stops at the first bad row). Auto-detects arbitrary
    (non-template) layouts; when the columns can't be recognized, returns
    ``confident: False`` with the detected columns instead of raising, so the UI
    can steer the user to the template or a manual mapping.
    """
    try:
        bound = max(0, int(limit))
    except (TypeError, ValueError):
        bound = 200
    rows = _read_ledger_rows(file_path)
    header_row = next((row for row in rows if any(str_or_none(cell) for cell in row)), None)
    detected = None
    native = column_map is None and header_row is not None and _is_native_ledger(header_row)
    if not native:
        header = [str(cell).strip() if cell is not None else "" for cell in (header_row or [])]
        if column_map is not None:
            plan, detected = column_map, None
        else:
            inferred = infer_ledger_columns(header)
            plan, detected = inferred["plan"], inferred["detected"]
        if not _ledger_plan_usable(plan):
            data_rows = [
                raw for raw in rows[(rows.index(header_row) + 1 if header_row is not None else 0):]
                if any(str_or_none(cell) for cell in raw)
            ]
            return {
                "confident": False,
                "detected": detected,
                "headers": header,
                "rows_read": len(data_rows),
                "mapped": 0,
                "errors": 0,
                "problems": [],
                "preview": [],
                "truncated": False,
            }
    records, _ = _ledger_source_records(rows, column_map)
    normalized = []
    problems = []
    for index, record in enumerate(records, start=1):
        try:
            normalized.append(normalize_generic_ledger_record(record, index=index))
        except AppError as exc:
            problems.append({"row": index, "message": str(exc)})
    return {
        "rows_read": len(records),
        "mapped": len(normalized),
        "errors": len(problems),
        "problems": problems[:bound] if bound else problems,
        "preview": [_generic_ledger_preview_row(record) for record in normalized[:bound]],
        "truncated": bound > 0 and len(normalized) > bound,
        "confident": True,
        "detected": detected,
    }


def is_generic_ledger_format(input_format):
    return input_format == GENERIC_LEDGER_FORMAT


# Bitcoin-only example rows for the fill-in template. Never altcoins.
_GENERIC_LEDGER_EXAMPLE_ROWS = (
    ("Buy", "2026-01-15", "0.05000000", "BTC", "3000.00", "EUR", "3.50", "EUR", "", "Coinfinity", "First stack", ""),
    ("Sell", "2026-02-10", "2200.00", "EUR", "0.03000000", "BTC", "1.00", "EUR", "", "Kraken", "Took some profit", ""),
    ("Mining", "2026-03-10", "0.00050000", "BTC", "", "", "", "", "32.50", "Solo pool", "Block reward", ""),
    ("Income", "2026-03-20", "250000", "SATS", "", "", "", "", "160.00", "Freelance client", "Invoice paid in sats", ""),
    ("Withdrawal", "2026-04-01", "", "", "0.02000000", "BTC", "0.00002000", "BTC", "", "", "Moved to cold storage", ""),
    ("Spend", "2026-04-15", "", "", "0.00100000", "BTC", "", "", "65.00", "Local merchant", "Coffee and groceries", ""),
    ("Gift sent", "2026-05-01", "", "", "0.00100000", "BTC", "", "", "", "", "Birthday gift to a friend", ""),
)

_GENERIC_LEDGER_LEGEND = (
    ("How to fill in this ledger", ""),
    ("", ""),
    ("One row per transaction. Each row has exactly one Bitcoin leg.", ""),
    ("Type", "What happened. Pick from the dropdown / list below."),
    ("Date", "Required. 2026-01-15, a full timestamp, or 15.01.2026."),
    ("Received Amount / Asset", "What came in. For a Buy: the Bitcoin you bought."),
    ("Sent Amount / Asset", "What went out. For a Buy: the fiat you paid."),
    ("Fee Amount / Asset", "Optional. Blank Fee Asset means the fee is in Bitcoin."),
    ("", "A fiat fee must match the row's fiat currency."),
    ("Fiat Value", "Fair-market value in your book currency. Use it for"),
    ("", "Income/Mining/Spend/Gift rows that have no cash leg."),
    ("Counterparty", "Optional. Exchange, merchant, or person."),
    ("Note", "Optional free text. Stored on the transaction."),
    ("Tx-ID", "Optional but recommended — lets you re-import safely."),
    ("", ""),
    ("Amounts are in BTC (e.g. 0.05000000) unless the Asset is SATS.", ""),
    ("A comma or dot decimal both work (0,05 = 0.05).", ""),
    ("Fiat columns must be in your book's currency (a EUR book takes", ""),
    ("EUR; a JPY book takes JPY). Mixed currencies are rejected.", ""),
    ("Gift sent / Donation / Lost / Stolen are flagged for review,", ""),
    ("not auto-sold at market value.", ""),
    ("To move Bitcoin between your own wallets, import a Withdrawal", ""),
    ("into the source wallet and a Deposit into the destination", ""),
    ("wallet with the same Tx-ID.", ""),
)


def _generic_ledger_template_format(output_path):
    return "csv" if str(output_path).lower().endswith(".csv") else "xlsx"


def _write_generic_ledger_csv_template(output_path):
    with open(output_path, "w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(GENERIC_LEDGER_COLUMNS)
        for row in _GENERIC_LEDGER_EXAMPLE_ROWS:
            writer.writerow(row)


def _write_generic_ledger_xlsx_template(output_path):
    try:
        import xlsxwriter
    except ImportError as exc:  # pragma: no cover - XlsxWriter is a hard dependency
        raise AppError(
            "Writing .xlsx templates needs the XlsxWriter package",
            code="dependency_missing",
        ) from exc
    workbook = xlsxwriter.Workbook(output_path, {"in_memory": True})
    try:
        header_fmt = workbook.add_format(
            {"bold": True, "bg_color": "#F2A900", "font_color": "#1A1A1A", "border": 1}
        )
        example_fmt = workbook.add_format({"italic": True, "font_color": "#666666"})
        title_fmt = workbook.add_format({"bold": True, "font_size": 13})
        term_fmt = workbook.add_format({"bold": True})

        sheet = workbook.add_worksheet("Transactions")
        sheet.freeze_panes(1, 0)
        for column_index, name in enumerate(GENERIC_LEDGER_COLUMNS):
            sheet.write(0, column_index, name, header_fmt)
        for row_index, row in enumerate(_GENERIC_LEDGER_EXAMPLE_ROWS, start=1):
            for column_index, value in enumerate(row):
                sheet.write(row_index, column_index, value, example_fmt)
        sheet.set_column(0, 0, 14)
        sheet.set_column(1, 1, 12)
        sheet.set_column(2, 7, 14)
        sheet.set_column(8, 8, 12)
        sheet.set_column(9, 11, 22)
        last_row = max(200, len(_GENERIC_LEDGER_EXAMPLE_ROWS) + 50)
        sheet.data_validation(
            1,
            0,
            last_row,
            0,
            {
                "validate": "list",
                "source": list(GENERIC_LEDGER_DISPLAY_TYPES),
                "error_message": "Pick a Type from the list (see the Legend sheet).",
            },
        )

        legend = workbook.add_worksheet("Legend")
        legend.set_column(0, 0, 32)
        legend.set_column(1, 1, 60)
        legend.write(0, 0, "Generic ledger import", title_fmt)
        for row_index, (term, definition) in enumerate(_GENERIC_LEDGER_LEGEND, start=2):
            legend.write(row_index, 0, term, term_fmt if definition else None)
            legend.write(row_index, 1, definition)
        type_start = len(_GENERIC_LEDGER_LEGEND) + 4
        legend.write(type_start, 0, "Transaction Types", title_fmt)
        cursor = type_start + 1
        for group_label, labels in GENERIC_LEDGER_TYPE_GROUPS:
            legend.write(cursor, 0, group_label, term_fmt)
            legend.write(cursor, 1, ", ".join(labels))
            cursor += 1
    finally:
        workbook.close()


def write_generic_ledger_template(output_path, fmt=None):
    """Write the fill-in ledger template to `output_path` (.xlsx or .csv)."""
    resolved = (fmt or _generic_ledger_template_format(output_path)).lower()
    if resolved == "csv":
        _write_generic_ledger_csv_template(output_path)
    elif resolved in {"xlsx", "xlsm"}:
        _write_generic_ledger_xlsx_template(output_path)
        resolved = "xlsx"
    else:
        raise AppError(
            f"Unsupported template format '{resolved}'",
            code="validation",
            hint="Use xlsx or csv.",
        )
    return {
        "file": os.path.abspath(output_path),
        "format": resolved,
        "columns": list(GENERIC_LEDGER_COLUMNS),
        "types": list(GENERIC_LEDGER_DISPLAY_TYPES),
    }


# -- BIP329 ------------------------------------------------------------------


def normalize_bip329_record(record):
    """Validate and normalize a single BIP329 label record.

    Enforces the known `record_type` set, non-empty `ref`, and the
    `spendable`-is-`output`-only rule from the spec.
    """
    if not isinstance(record, dict):
        raise AppError("BIP329 records must be JSON objects")
    record_type = str(record.get("type") or "").strip()
    ref = str(record.get("ref") or "").strip()
    if record_type not in {"tx", "addr", "pubkey", "input", "output", "xpub", "spscan"}:
        raise AppError(f"Unsupported BIP329 record type '{record_type}'")
    if not ref:
        raise AppError("BIP329 records require a non-empty ref")
    spendable = record.get("spendable")
    if spendable is not None and not isinstance(spendable, bool):
        raise AppError("BIP329 spendable must be a boolean when present")
    if spendable is not None and record_type != "output":
        raise AppError("BIP329 spendable is only valid for output records")
    return {
        "type": record_type,
        "ref": ref,
        "label": str_or_none(record.get("label")),
        "origin": str_or_none(record.get("origin")),
        "spendable": spendable,
        "data": {
            key: value
            for key, value in record.items()
            if key not in {"type", "ref", "label", "origin", "spendable"}
        },
    }


def load_bip329_file(file_path):
    """Load a BIP329 JSONL file (one JSON object per line) and normalize each row."""
    records = []
    with open(file_path, "r", encoding="utf-8") as handle:
        for line_number, raw_line in enumerate(handle, start=1):
            line = raw_line.strip()
            if not line:
                continue
            try:
                payload = json.loads(line)
            except json.JSONDecodeError as exc:
                raise AppError(f"Invalid BIP329 JSON on line {line_number}") from exc
            records.append(normalize_bip329_record(payload))
    return records
