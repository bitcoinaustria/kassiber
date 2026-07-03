from __future__ import annotations

import contextlib
import random
import socket
import tempfile
import unittest
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

import openpyxl

from kassiber import backends as backend_config
from kassiber.core import accounts as core_accounts
from kassiber.core import imports as core_imports
from kassiber.core import output_inventory as core_output_inventory
from kassiber.core import rates as core_rates
from kassiber.core import reports as core_reports
from kassiber.core import sync as core_sync
from kassiber.core import sync_backends as core_sync_backends
from kassiber.core import wallets as core_wallets
from kassiber.core.repo import invalidate_journals
from kassiber.cli.handlers import _report_hooks as cli_report_hooks
from kassiber.cli.handlers import process_journals
from kassiber.db import open_db

from tests.integration.env import no_egress_guard
from tests.integration import regtest_demo
from tests.integration.tapes import BitcoinRpcTape, RecordedTape, TapeMiss


ROOT = Path(__file__).resolve().parent.parent
TAPE = ROOT / "tests" / "fixtures" / "regtest_tapes" / "bitcoin_core_address_baseline.json"
EDGE_TAPE = ROOT / "tests" / "fixtures" / "regtest_tapes" / "bitcoin_core_edge_cases.json"


_EGRESS_STACK: contextlib.ExitStack | None = None


def setUpModule() -> None:
    """Install the no-egress guard for the whole fast lane when requested.

    `integration-harness.sh fast` exports `KASSIBER_NO_EGRESS=1`; passing
    `enabled=None` makes the guard consult that flag, so the entire replay lane
    blocks non-loopback connects by default instead of relying on each test to
    opt in. Under the plain quality gate (flag unset) this is a no-op, and the
    per-test `no_egress_guard(enabled=True)` blocks still protect the critical
    replay paths there.
    """
    global _EGRESS_STACK
    _EGRESS_STACK = contextlib.ExitStack()
    _EGRESS_STACK.enter_context(no_egress_guard())


def tearDownModule() -> None:
    global _EGRESS_STACK
    if _EGRESS_STACK is not None:
        _EGRESS_STACK.close()
        _EGRESS_STACK = None
REGTEST_ADDRESS = "bcrt1qs758ursh4q9z627kt3pp5yysm78ddny6txaqgw"
EDGE_ADDRESSES = [
    "bcrt1qv4jxwefdvdshxefdv9jxgu3ddahx2gfplsfksc",
    "bcrt1qv4jxwefdvdshxefdv9jxgu3dw3mk7gfpl9er9m",
]


def _import_hooks() -> core_imports.ImportCoordinatorHooks:
    return core_imports.ImportCoordinatorHooks(
        ensure_tag_row=lambda *args: None,
        invalidate_journals=invalidate_journals,
    )


def _sync_hooks() -> core_sync.WalletSyncHooks:
    return core_sync.WalletSyncHooks(
        import_file=lambda *args: {},
        insert_records=lambda db, prof, wal, records, source_label: core_imports.insert_wallet_records(
            db,
            prof,
            wal,
            records,
            source_label,
            _import_hooks(),
        ),
        resolve_backend=backend_config.resolve_backend,
        resolve_sync_state=core_sync_backends.resolve_wallet_sync_targets,
        normalize_addresses=core_wallets.normalize_addresses,
        backend_adapters={"bitcoinrpc": core_sync_backends.bitcoinrpc_sync_adapter},
        update_output_inventory=core_output_inventory.update_wallet_output_inventory,
    )


def _create_replay_book(conn, data_root: Path, *, core_wallet: str, addresses: list[str]):
    workspace = core_accounts.create_workspace(conn, "Regtest")
    profile = core_accounts.create_profile(
        conn,
        workspace["id"],
        "Replay",
        "EUR",
        "FIFO",
        "generic",
        365,
    )
    core_accounts.create_backend(
        conn,
        "core-regtest",
        "bitcoinrpc",
        "http://127.0.0.1:18443",
        chain="bitcoin",
        network="regtest",
        timeout=30,
        config={"wallet": core_wallet, "username": "user", "password": "pass"},
    )
    wallet = core_wallets.create_wallet(
        conn,
        workspace["id"],
        profile["id"],
        "Core replay",
        "address",
        None,
        {
            "backend": "core-regtest",
            "chain": "bitcoin",
            "network": "regtest",
            "addresses": list(addresses),
        },
    )
    runtime_config = backend_config.merge_db_backends(
        conn,
        {
            "env_file": str(data_root / "unused.env"),
            "default_backend": "core-regtest",
            "bootstrap_default_backend": "core-regtest",
            "backends": {},
            "bootstrap_backends": {},
            "dotenv_backends": [],
            "process_env_overrides": {"backends": {}, "default_backend": False},
        },
    )
    profile_row = conn.execute("SELECT * FROM profiles WHERE id = ?", (profile["id"],)).fetchone()
    wallet_row = conn.execute("SELECT * FROM wallets WHERE id = ?", (wallet["id"],)).fetchone()
    return workspace, profile, wallet, profile_row, wallet_row, runtime_config


class RegtestHarnessTest(unittest.TestCase):
    def test_no_egress_guard_blocks_only_non_loopback_connects(self):
        calls = []

        def fake_connect(_self, address):
            calls.append(("connect", address))
            return None

        def fake_connect_ex(_self, address):
            calls.append(("connect_ex", address))
            return 0

        with patch("socket.socket.connect", fake_connect), patch(
            "socket.socket.connect_ex",
            fake_connect_ex,
        ):
            with no_egress_guard(enabled=True):
                with self.assertRaises(AssertionError):
                    socket.socket.connect(object(), ("198.51.100.1", 443))
                socket.socket.connect(object(), ("127.0.0.1", 18443))
                socket.socket.connect_ex(object(), ("localhost", 18443))

        self.assertEqual(
            calls,
            [
                ("connect", ("127.0.0.1", 18443)),
                ("connect_ex", ("localhost", 18443)),
            ],
        )

    def test_recorded_tape_has_provenance_and_completeness_gate(self):
        tape = RecordedTape.load(TAPE)

        self.assertEqual(tape.provenance["kassiber_issue"], 312)
        self.assertEqual(tape.provenance["backend_kind"], "bitcoinrpc")
        with self.assertRaises(TapeMiss):
            tape.lookup("missing interaction")

    def test_full_accounting_demo_manifest_covers_expected_workflows(self):
        scenario = regtest_demo.load_scenario()

        self.assertEqual(scenario["id"], "full-accounting-v1")
        self.assertEqual(scenario["backend"]["name"], "core-regtest")
        self.assertEqual(scenario["backend"]["default"], "bitcoin-mempool-regtest")
        wallet_keys = {wallet["key"] for wallet in scenario["wallets"]}
        self.assertEqual(
            wallet_keys,
            {
                "treasury",
                "cold",
                "spending",
                "merchant",
                "silent_payments",
                "miner",
                "treasury_2020",
                "merchant_2022",
                "cold_2024",
                "liquid_treasury",
                "liquid_operations",
                "liquid_treasury_2024",
                "liquid_live_sync",
            },
        )
        liquid_wallets = {wallet["key"] for wallet in scenario["wallets"] if wallet.get("chain") == "liquid"}
        self.assertEqual(
            liquid_wallets,
            {"liquid_treasury", "liquid_operations", "liquid_treasury_2024", "liquid_live_sync"},
        )
        self.assertEqual(
            {wallet["network"] for wallet in scenario["wallets"] if wallet.get("chain") == "liquid"},
            {"elementsregtest"},
        )
        local_backends = regtest_demo._local_backend_specs()
        self.assertIn(scenario["backend"]["default"], {backend["name"] for backend in local_backends})
        self.assertEqual(
            {backend["name"] for backend in local_backends},
            {
                "bitcoin-electrum-regtest",
                "bitcoin-frigate-regtest",
                "bitcoin-mempool-regtest",
                "liquid-electrum-regtest",
                "liquid-mempool-regtest",
            },
        )
        self.assertTrue(all("127.0.0.1" in backend["url"] for backend in local_backends))
        self.assertEqual(
            {(backend["chain"], backend["network"]) for backend in local_backends},
            {("bitcoin", "regtest"), ("liquid", "elementsregtest")},
        )
        boltz_bridges = [
            bridge for bridge in scenario["stress"]["swap_bridges"] if bridge.get("provider") == "boltz"
        ]
        self.assertEqual(len(boltz_bridges), 1)
        self.assertEqual(
            {
                (
                    bridge["boltz_flow"],
                    bridge["boltz_api"],
                    bridge["boltz_from"],
                    bridge["boltz_to"],
                )
                for bridge in boltz_bridges
            },
            {("chain-swap", "/v2/swap/chain", "BTC", "L-BTC")},
        )
        self.assertEqual(
            scenario["deprecated_wallets"],
            ["treasury", "merchant", "cold", "liquid_treasury"],
        )
        # Every operational Bitcoin wallet rotates through several watched
        # addresses so the demo book looks like real wallet usage, not one
        # reused address per wallet.
        multi_address_wallets = {
            wallet["key"]
            for wallet in scenario["wallets"]
            if int(wallet.get("addresses") or 1) > 1
        }
        self.assertTrue(
            {"treasury", "cold", "spending", "merchant", "miner"}.issubset(multi_address_wallets)
        )
        silent_payment_wallet = next(wallet for wallet in scenario["wallets"] if wallet["key"] == "silent_payments")
        self.assertEqual(silent_payment_wallet["kind"], "silent-payment")
        self.assertEqual(silent_payment_wallet["network"], "regtest")
        self.assertTrue(str(silent_payment_wallet["sp_descriptor"]).startswith("sp(tspscan1q"))
        operation_kinds = {operation["kind"] for operation in scenario["operations"]}
        self.assertTrue(
            {
                "payment",
                "self_transfer",
                "coinjoin_shape",
                "payjoin_shape",
                "rbf_replaced_payment",
                "external_receipt",
                "batched_payment",
                "incoming_burst",
                "many_input_consolidation",
                "loan_collateral_lock",
                "loan_collateral_release",
                "loan_principal_received",
                "loan_principal_repaid",
            }.issubset(operation_kinds)
        )
        dust_amounts = [
            Decimal(operation["amount_btc"])
            for operation in scenario["operations"]
            if operation["kind"] == "external_receipt"
        ]
        self.assertTrue(any(amount < Decimal("0.00001") for amount in dust_amounts))
        pending = scenario["pending_operations"]
        self.assertEqual([op["kind"] for op in pending], ["external_receipt"])
        self.assertEqual(scenario["expected"]["wallets"], 13)
        self.assertEqual(scenario["expected"]["deprecated_wallets"], 4)
        self.assertEqual(scenario["expected"]["assets"], ["BTC", "LBTC"])
        self.assertGreaterEqual(scenario["expected"]["min_transactions"], 847)
        self.assertGreaterEqual(scenario["expected"]["min_active_transactions"], 842)
        self.assertEqual(scenario["expected"]["pending_transactions"], 1)
        base_time = datetime.fromisoformat(scenario["base_time"].replace("Z", "+00:00"))
        stress = scenario["stress"]
        self.assertTrue(stress["enabled"])
        self.assertGreaterEqual(stress["cycles"], 130)
        stress_span_days = stress["cycles"] * stress["days_between_cycles"]
        self.assertGreaterEqual(stress_span_days, 365 * 7)
        self.assertLess(base_time, datetime(2020, 1, 1, tzinfo=timezone.utc))
        latest_time = datetime.fromisoformat(scenario["latest_time"].replace("Z", "+00:00"))
        estimated_end = datetime.fromtimestamp(
            regtest_demo.estimate_scenario_end_ts(scenario),
            tz=timezone.utc,
        )
        self.assertLessEqual(estimated_end, latest_time)
        self.assertLess(latest_time, datetime(2026, 7, 3, tzinfo=timezone.utc))
        self.assertTrue(stress["business_expenses"]["enabled"])
        self.assertGreaterEqual(len(stress["business_expenses"]["schedule"]), 6)
        self.assertEqual(len(stress["wallet_rotations"]), 3)
        self.assertEqual(len(stress["swap_bridges"]), 3)
        # Deterministic amount/fee variation keeps the multi-year ledger from
        # looking like a spreadsheet of identical round numbers.
        self.assertGreaterEqual(int(stress["variation_bp"]), 1000)
        mining_events = stress["mining_events"]
        self.assertEqual([event["role"] for event in mining_events], ["miner", "miner"])
        for event in mining_events:
            # Coinbase rewards must have >= 100 blocks left to mature before sync.
            self.assertLessEqual(int(event["cycle"]), stress["cycles"] - 35)
        liquid_rows = sum(len(rows) for rows in scenario["liquid_ledger"]["wallets"].values())
        self.assertGreaterEqual(liquid_rows, 9)
        self.assertEqual(len(scenario["liquid_ledger"]["transfer_pairs"]), 1)
        liquid_live_wallets = [
            wallet for wallet in scenario["wallets"]
            if wallet.get("chain") == "liquid" and wallet.get("kind") == "descriptor"
        ]
        self.assertEqual([wallet["key"] for wallet in liquid_live_wallets], ["liquid_live_sync"])
        self.assertGreater(Decimal(liquid_live_wallets[0]["live_receipt_btc"]), Decimal("0"))
        self.assertEqual(scenario["pricing"]["source"], "kraken-bundled")
        self.assertEqual(scenario["pricing"]["live_source"], "mempool")
        self.assertEqual(scenario["expected"]["pricing_source"], "kraken-csv")
        self.assertIn("LBTC", scenario["expected"]["require_pricing_provider_assets"])
        rates = [float(rate) for rate in scenario["pricing"]["fallback"]["rate_sequence"]]
        self.assertGreater(max(rates) / min(rates), 2.0)
        self.assertNotEqual(rates, sorted(rates))
        self.assertNotEqual(rates, sorted(rates, reverse=True))
        self.assertEqual(scenario["expected"]["collaborative_excluded"], 5)
        self.assertEqual(scenario["expected"]["min_transfer_pairs"], 9)
        self.assertEqual(scenario["expected"]["loan_marks"], 4)
        self.assertIn("full-report.xlsx", scenario["expected"]["export_files"])

    def test_full_accounting_demo_manifest_validation_fails_closed(self):
        scenario = regtest_demo.load_scenario()
        scenario.pop("operations")

        with self.assertRaisesRegex(ValueError, "operations"):
            regtest_demo.validate_scenario(scenario)

    def test_compose_stack_includes_local_protocol_backends(self):
        compose = (ROOT / "dev" / "regtest" / "compose.bitcoin.yml").read_text(encoding="utf-8")

        self.assertIn("elementsd:", compose)
        self.assertIn("fulcrum:", compose)
        self.assertIn("bitcoin-mempool:", compose)
        self.assertIn("liquid-electrum:", compose)
        self.assertIn("liquid-mempool:", compose)
        self.assertIn("frigate:", compose)
        self.assertIn("profiles:", compose)
        self.assertIn("silent-payments", compose)
        self.assertIn("bitcoin/bitcoin:30.0", compose)
        self.assertIn("zmqpubsequence=tcp://0.0.0.0:28336", compose)
        self.assertIn("Dockerfile.frigate", compose)
        self.assertIn("backendElectrumServer = \"tcp://fulcrum:50001\"", compose)
        self.assertNotIn("backend-stack:", compose)
        self.assertIn("KASSIBER_REGTEST_ELEMENTSD_IMAGE", compose)
        self.assertIn("KASSIBER_REGTEST_FULCRUM_IMAGE", compose)
        self.assertIn("KASSIBER_REGTEST_ELEMENTS_RPC_PORT", compose)
        self.assertIn("backend_stack.py:/app/backend_stack.py:ro", compose)
        self.assertIn("KASSIBER_REGTEST_BITCOIN_ELECTRUM_PORT", compose)
        self.assertIn("KASSIBER_REGTEST_BITCOIN_MEMPOOL_PORT", compose)
        self.assertIn("KASSIBER_REGTEST_LIQUID_ELECTRUM_PORT", compose)
        self.assertIn("KASSIBER_REGTEST_LIQUID_MEMPOOL_PORT", compose)
        self.assertIn("KASSIBER_REGTEST_FRIGATE_PORT", compose)
        self.assertIn("host.docker.internal:host-gateway", compose)

    def test_local_backend_stack_exposes_mempool_price_api(self):
        backend_stack = (ROOT / "dev" / "regtest" / "backend_stack.py").read_text(encoding="utf-8")

        self.assertIn("/api/v1/prices", backend_stack)
        self.assertIn("/api/v1/historical-price", backend_stack)
        self.assertIn("KASSIBER_REGTEST_BTC_EUR_PRICE", backend_stack)
        self.assertIn("do_OPTIONS", backend_stack)
        self.assertIn("Access-Control-Allow-Origin", backend_stack)
        self.assertIn("KASSIBER_REGTEST_EXPLORER_CORS_ORIGIN", backend_stack)

    def test_demo_latest_rate_seed_does_not_change_book_provider(self):
        demo = (ROOT / "tests" / "integration" / "regtest_demo.py").read_text(encoding="utf-8")
        harness = (ROOT / "scripts" / "integration-harness.sh").read_text(encoding="utf-8")

        self.assertNotIn("set_market_rate_provider", demo)
        self.assertNotIn("set_market_rate_provider", harness)
        self.assertIn("sync_latest_rates", demo)
        self.assertIn("sync_latest_rates", harness)

    def test_demo_up_keeps_interactive_book_report_ready(self):
        harness = (ROOT / "scripts" / "integration-harness.sh").read_text(encoding="utf-8")

        self.assertIn("--no-business-tick", harness)
        self.assertIn("demo_load_rpc_env", harness)
        self.assertIn("KASSIBER_REGTEST_COMPOSE_PROFILES", harness)

    def test_silent_payments_lane_probes_frigate(self):
        harness = (ROOT / "scripts" / "integration-harness.sh").read_text(encoding="utf-8")

        self.assertIn("run_silent_payments()", harness)
        self.assertIn("wait_for_frigate", harness)
        self.assertIn("seed_frigate_regtest_tip", harness)
        self.assertIn("generatetoaddress", harness)
        self.assertIn("server.features", harness)
        self.assertIn("silent-payments", harness)
        self.assertIn("KASSIBER_REGTEST_FRIGATE_PORT", harness)
        self.assertIn("KASSIBER_REGTEST_FRIGATE_WAIT_SECONDS", harness)

    def test_demo_purge_paths_are_guarded_by_safe_home_check(self):
        harness = (ROOT / "scripts" / "integration-harness.sh").read_text(encoding="utf-8")

        self.assertIn("demo_assert_safe_home()", harness)
        self.assertIn("path is root, user home, a temp root, or root-level", harness)
        self.assertIn("path is too shallow", harness)
        self.assertIn("missing Kassiber regtest demo manifest", harness)
        self.assertLess(
            harness.index("demo_assert_safe_home rebuild"),
            harness.index('rm -rf "$DEMO_HOME/data"'),
        )
        self.assertLess(
            harness.index("demo_assert_safe_home purge"),
            harness.index('rm -rf "$DEMO_HOME"'),
        )

    def test_demo_manifest_writer_uses_private_atomic_replacement(self):
        harness = (ROOT / "scripts" / "integration-harness.sh").read_text(encoding="utf-8")

        self.assertIn("os.chmod(home, 0o700)", harness)
        self.assertIn("os.open(candidate, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600)", harness)
        self.assertIn("os.replace(tmp_path, manifest_path)", harness)
        self.assertNotIn('open(manifest_path, "w"', harness)

    def test_full_accounting_demo_manifest_validation_rejects_bad_edge_cases(self):
        scenario = regtest_demo.load_scenario()
        rbf = next(op for op in scenario["operations"] if op["kind"] == "rbf_replaced_payment")
        rbf_fee = rbf["replacement_fee_btc"]
        rbf["replacement_fee_btc"] = rbf["fee_btc"]
        with self.assertRaisesRegex(ValueError, "replacement fee"):
            regtest_demo.validate_scenario(scenario)
        rbf["replacement_fee_btc"] = rbf_fee

        pending = scenario["pending_operations"][0]
        pending_kind = pending["kind"]
        pending["kind"] = "payment"
        with self.assertRaisesRegex(ValueError, "external_receipt"):
            regtest_demo.validate_scenario(scenario)
        pending["kind"] = pending_kind

        mining_event = scenario["stress"]["mining_events"][0]
        mining_cycle = mining_event["cycle"]
        mining_event["cycle"] = scenario["stress"]["cycles"]
        with self.assertRaisesRegex(ValueError, "too late to mature"):
            regtest_demo.validate_scenario(scenario)
        mining_event["cycle"] = mining_cycle

        scenario["wallets"][0]["addresses"] = 99
        with self.assertRaisesRegex(ValueError, "addresses"):
            regtest_demo.validate_scenario(scenario)

    def test_core_rpc_tape_replays_through_sync_journal_report_and_xlsx(self):
        tape_rpc = BitcoinRpcTape(RecordedTape.load(TAPE))
        with tempfile.TemporaryDirectory() as tmp:
            # resolve() so path assertions survive macOS /var -> /private/var symlinks
            data_root = Path(tmp).resolve() / "data"
            export_path = Path(tmp).resolve() / "report.xlsx"
            conn = open_db(data_root)
            try:
                workspace, profile, wallet, profile_row, wallet_row, runtime_config = _create_replay_book(
                    conn,
                    data_root,
                    core_wallet="kassiber-wallet-1",
                    addresses=[REGTEST_ADDRESS],
                )
                hooks = _sync_hooks()

                with no_egress_guard(enabled=True), patch(
                    "kassiber.core.sync_backends.bitcoinrpc_call",
                    tape_rpc.call,
                ):
                    outcome = core_sync.sync_wallet_from_backend(
                        conn,
                        runtime_config,
                        profile_row,
                        wallet_row,
                        hooks,
                    )
                    self.assertEqual(outcome["backend_kind"], "bitcoinrpc")
                    self.assertEqual(outcome["records_fetched"], 2)
                    self.assertEqual(outcome["imported"], 2)
                    self.assertEqual(outcome["bitcoinrpc_sync_mode"], "full_scan")
                    self.assertEqual(outcome["output_inventory"]["observed"], 1)
                    self.assertEqual(outcome["output_inventory"]["active"], 1)
                    self.assertIn("bitcoinrpc_last_block", outcome["freshness_checkpoint"])

                    tx_rows = conn.execute(
                        """
                        SELECT external_id, direction, amount, fee
                        FROM transactions
                        ORDER BY occurred_at
                        """
                    ).fetchall()
                    self.assertEqual(len(tx_rows), 2)
                    self.assertEqual(tx_rows[0]["direction"], "inbound")
                    self.assertEqual(tx_rows[0]["amount"], 25_000_000_000)
                    self.assertEqual(tx_rows[1]["direction"], "outbound")
                    self.assertEqual(tx_rows[1]["amount"], 5_000_000_000)
                    self.assertEqual(tx_rows[1]["fee"], 1_000_000)

                    utxo = conn.execute(
                        """
                        SELECT wu.txid, wu.vout, wu.amount, wu.block_height, wu.block_time, t.id AS transaction_id
                        FROM wallet_utxos wu
                        LEFT JOIN transactions t
                          ON t.profile_id = wu.profile_id
                         AND t.external_id = wu.txid
                        """
                    ).fetchone()
                    self.assertIsNotNone(utxo)
                    self.assertEqual(
                        utxo["txid"],
                        "bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb",
                    )
                    self.assertEqual(utxo["vout"], 0)
                    self.assertEqual(utxo["amount"], 19_999_000_000)
                    self.assertEqual(utxo["block_height"], 102)
                    self.assertEqual(utxo["block_time"], "2023-11-14T22:23:20Z")
                    self.assertIsNotNone(utxo["transaction_id"])

                    core_rates.set_manual_rate(conn, "BTC-EUR", "2023-11-14T22:13:20Z", "35000")
                    core_rates.set_manual_rate(conn, "BTC-EUR", "2023-11-14T22:23:20Z", "36000")

                    report_hooks = cli_report_hooks()
                    journal = process_journals(conn, workspace["id"], profile["id"])
                    self.assertEqual(journal["entries_created"], 2)
                    self.assertEqual(journal["quarantined"], 0)
                    self.assertEqual(journal["auto_priced"], 2)
                    self.assertEqual(journal["processed_transactions"], 2)

                    priced = conn.execute(
                        """
                        SELECT external_id, fiat_rate, fiat_value, pricing_source_kind, pricing_quality
                        FROM transactions
                        ORDER BY occurred_at
                        """
                    ).fetchall()
                    self.assertEqual([row["fiat_value"] for row in priced], [8750.0, 1800.0])
                    self.assertEqual({row["pricing_source_kind"] for row in priced}, {"manual_rate_cache"})
                    self.assertEqual({row["pricing_quality"] for row in priced}, {"exact"})

                    journal_rows = conn.execute(
                        """
                        SELECT entry_type, asset, quantity, fiat_value, cost_basis, proceeds, gain_loss
                        FROM journal_entries
                        ORDER BY occurred_at, entry_type
                        """
                    ).fetchall()
                    self.assertEqual(len(journal_rows), 2)
                    self.assertEqual(journal_rows[0]["entry_type"], "acquisition")
                    self.assertEqual(journal_rows[0]["quantity"], 25_000_000_000)
                    self.assertEqual(journal_rows[0]["fiat_value"], 8750.0)
                    self.assertEqual(journal_rows[1]["entry_type"], "disposal")
                    self.assertEqual(journal_rows[1]["quantity"], -5_001_000_000)
                    self.assertEqual(journal_rows[1]["proceeds"], 1800.0)
                    self.assertAlmostEqual(journal_rows[1]["cost_basis"], 1750.35, places=2)
                    self.assertAlmostEqual(journal_rows[1]["gain_loss"], 49.65, places=2)

                    summary = core_reports.report_summary(conn, workspace["id"], profile["id"], report_hooks)
                    self.assertEqual(summary["metrics"]["assets_in_scope"], 1)
                    self.assertEqual(summary["metrics"]["journal_entries"], 2)
                    self.assertEqual(summary["metrics"]["priced_transactions"], 2)
                    self.assertEqual(summary["metrics"]["quarantines"], 0)
                    self.assertAlmostEqual(summary["holdings"]["cost_basis"], 6999.65, places=2)
                    self.assertAlmostEqual(summary["realized"]["gain_loss"], 49.65, places=2)

                    export = core_reports.export_xlsx_report(
                        conn,
                        workspace["id"],
                        profile["id"],
                        str(export_path),
                        report_hooks,
                        verify=True,
                    )
                    self.assertEqual(export["file"], str(export_path))
                    self.assertTrue(export_path.exists())
                    self.assertTrue(export["verified"])

                    workbook = openpyxl.load_workbook(export_path, data_only=False, read_only=True)
                    self.assertIn("Verify", workbook.sheetnames)
                    self.assertIn("Control", workbook.sheetnames)
                    self.assertIn("Acquisitions", workbook.sheetnames)
                    self.assertIn("Disposals", workbook.sheetnames)
                    self.assertEqual(workbook["Verify"]["A2"].value, "Verification status")
                    self.assertIn("ALL CHECKS OK", workbook["Verify"]["B2"].value)
                    self.assertEqual(workbook["Acquisitions"]["C3"].value, tx_rows[0]["external_id"])
                    self.assertEqual(workbook["Acquisitions"]["F3"].value, 25_000_000_000)
                    self.assertEqual(workbook["Acquisitions"]["H3"].value, 8750)
                    self.assertEqual(workbook["Disposals"]["C3"].value, tx_rows[1]["external_id"])
                    self.assertEqual(workbook["Disposals"]["F3"].value, 5_001_000_000)
                    self.assertEqual(workbook["Disposals"]["K3"].value, 49.65)
                    self.assertIn('"OK"', workbook["Disposals"]["L3"].value)
                    self.assertEqual(workbook["Control"]["A3"].value, "BTC")
                    self.assertEqual(workbook["Control"]["F3"].value, 0.19999)
                    self.assertEqual(workbook["Control"]["I3"].value, 6999.65)
                    self.assertIn('"OK"', workbook["Control"]["G3"].value)
                    workbook.close()

                    # Second run feeds the first checkpoint back into Core RPC and
                    # proves the incremental listsinceblock path plus DB idempotency.
                    repeat_wallet_row = conn.execute(
                        "SELECT * FROM wallets WHERE id = ?", (wallet["id"],)
                    ).fetchone()
                    repeat = core_sync.sync_wallet_from_backend(
                        conn,
                        runtime_config,
                        profile_row,
                        repeat_wallet_row,
                        hooks,
                        checkpoint=outcome["freshness_checkpoint"],
                    )
                self.assertEqual(repeat["imported"], 0)
                self.assertEqual(repeat["records_fetched"], 0)
                self.assertEqual(repeat["bitcoinrpc_sync_mode"], "sinceblock")
                count = conn.execute("SELECT COUNT(*) AS count FROM transactions").fetchone()["count"]
                self.assertEqual(count, 2)
                self.assertEqual(tape_rpc.unused_interactions(), [])
            finally:
                conn.close()

    def test_core_rpc_tape_replays_realistic_edge_case_wallet(self):
        # A messy but realistic wallet: two watched addresses, an immature and a
        # mature coinbase, a dust deposit, an RBF-replaced conflict pair, a
        # same-wallet self-spend, and a receipt still in the mempool at sync.
        tape_rpc = BitcoinRpcTape(RecordedTape.load(EDGE_TAPE))
        with tempfile.TemporaryDirectory() as tmp:
            data_root = Path(tmp) / "data"
            conn = open_db(data_root)
            try:
                workspace, profile, wallet, profile_row, wallet_row, runtime_config = _create_replay_book(
                    conn,
                    data_root,
                    core_wallet="kassiber-wallet-edge",
                    addresses=EDGE_ADDRESSES,
                )
                hooks = _sync_hooks()

                with no_egress_guard(enabled=True), patch(
                    "kassiber.core.sync_backends.bitcoinrpc_call",
                    tape_rpc.call,
                ):
                    outcome = core_sync.sync_wallet_from_backend(
                        conn,
                        runtime_config,
                        profile_row,
                        wallet_row,
                        hooks,
                    )

                self.assertEqual(outcome["backend_kind"], "bitcoinrpc")
                # 7 txids observed; the immature coinbase and the RBF-replaced
                # original must not become ledger rows.
                self.assertEqual(outcome["records_fetched"], 5)
                self.assertEqual(outcome["imported"], 5)
                self.assertEqual(outcome["output_inventory"]["observed"], 4)

                rows = {
                    row["external_id"]: row
                    for row in conn.execute(
                        """
                        SELECT external_id, direction, kind, amount, fee, occurred_at, confirmed_at
                        FROM transactions
                        """
                    ).fetchall()
                }
                self.assertNotIn("c1" * 32, rows)  # immature coinbase skipped
                self.assertNotIn("e1" * 32, rows)  # RBF-replaced original skipped

                mature_coinbase = rows["c2" * 32]
                self.assertEqual(mature_coinbase["direction"], "inbound")
                self.assertEqual(mature_coinbase["amount"], 2_500_000_000_000)

                dust = rows["d1" * 32]
                self.assertEqual(dust["direction"], "inbound")
                self.assertEqual(dust["amount"], 546_000)

                replacement = rows["e2" * 32]
                self.assertEqual(replacement["direction"], "outbound")
                self.assertEqual(replacement["amount"], 29_988_000_000)
                self.assertEqual(replacement["fee"], 12_000_000)

                self_spend = rows["f1" * 32]
                self.assertEqual(self_spend["direction"], "outbound")
                self.assertEqual(self_spend["kind"], "fee")
                self.assertEqual(self_spend["amount"], 0)
                self.assertEqual(self_spend["fee"], 1_800_000)

                pending = rows["a9" * 32]
                self.assertEqual(pending["direction"], "inbound")
                self.assertEqual(pending["amount"], 2_500_000_000)
                self.assertIsNone(pending["confirmed_at"])

                mempool_utxo = conn.execute(
                    "SELECT block_height, block_time FROM wallet_utxos WHERE txid = ?",
                    ("a9" * 32,),
                ).fetchone()
                self.assertIsNotNone(mempool_utxo)
                self.assertIsNone(mempool_utxo["block_height"])

                for occurred_at, rate in (
                    ("2023-11-14T22:13:20Z", "34000"),
                    ("2023-11-14T23:13:20Z", "34500"),
                    ("2023-11-15T00:13:20Z", "35000"),
                    ("2023-11-15T01:00:00Z", "35500"),
                    ("2023-11-15T02:00:00Z", "36000"),
                ):
                    core_rates.set_manual_rate(conn, "BTC-EUR", occurred_at, rate)

                journal = process_journals(conn, workspace["id"], profile["id"])
                self.assertEqual(journal["quarantined"], 0)
                self.assertEqual(journal["processed_transactions"], 5)
                self.assertEqual(journal["auto_priced"], 5)

                summary = core_reports.report_summary(
                    conn, workspace["id"], profile["id"], cli_report_hooks()
                )
                self.assertEqual(summary["metrics"]["quarantines"], 0)
                self.assertEqual(summary["metrics"]["active_transactions"], 5)

                self.assertEqual(tape_rpc.unused_interactions(), [])
            finally:
                conn.close()

    def test_full_accounting_demo_has_boom_and_bust_regimes(self):
        # Balances must not be monotonically up-and-to-the-right: the scenario
        # models both accumulation and drawdown periods.
        scenario = regtest_demo.load_scenario()
        regimes = scenario["stress"]["economic_regimes"]
        self.assertGreaterEqual(len(regimes), 4)
        self.assertLessEqual(
            sum(int(phase["cycles"]) for phase in regimes),
            scenario["stress"]["cycles"],
        )
        downturns = [p for p in regimes if Decimal(p["spend_scale"]) > Decimal(p["receipt_scale"])]
        booms = [p for p in regimes if Decimal(p["receipt_scale"]) > Decimal(p["spend_scale"])]
        self.assertTrue(downturns, "expected at least one drawdown regime")
        self.assertTrue(booms, "expected at least one boom regime")

    def test_regime_scales_select_phase_by_cycle(self):
        regimes = [
            {"label": "up", "cycles": 2, "receipt_scale": "1.6", "spend_scale": "0.7"},
            {"label": "down", "cycles": 3, "receipt_scale": "0.4", "spend_scale": "1.9"},
        ]
        self.assertEqual(regtest_demo._regime_scales(1, regimes)[2], "up")
        self.assertEqual(regtest_demo._regime_scales(2, regimes)[2], "up")
        self.assertEqual(regtest_demo._regime_scales(3, regimes)[2], "down")
        self.assertEqual(regtest_demo._regime_scales(5, regimes)[2], "down")
        # past the last phase -> neutral 1.0/1.0
        self.assertEqual(
            regtest_demo._regime_scales(99, regimes),
            (Decimal("1"), Decimal("1"), "steady"),
        )

    def test_manifest_validation_requires_boom_and_downturn(self):
        scenario = regtest_demo.load_scenario()
        scenario["stress"]["economic_regimes"] = [
            {"label": "only-up", "cycles": 5, "receipt_scale": "1.5", "spend_scale": "0.8"},
        ]
        with self.assertRaisesRegex(ValueError, "downturn"):
            regtest_demo.validate_scenario(scenario)

    def test_tick_plan_is_deterministic_and_well_formed(self):
        active = ["spending", "merchant_2022", "treasury_2020"]
        plan_a = regtest_demo.plan_tick_operations(
            active, random.Random(7), receipts=3, payments=2, transfers=1
        )
        plan_b = regtest_demo.plan_tick_operations(
            active, random.Random(7), receipts=3, payments=2, transfers=1
        )
        shape = lambda plan: [(op["kind"], op["wallet"], op["to"], str(op["amount_btc"])) for op in plan]
        self.assertEqual(shape(plan_a), shape(plan_b))
        self.assertEqual(sum(1 for op in plan_a if op["kind"] == "receipt"), 3)
        self.assertEqual(sum(1 for op in plan_a if op["kind"] == "payment"), 2)
        self.assertEqual(sum(1 for op in plan_a if op["kind"] == "transfer"), 1)
        for op in plan_a:
            self.assertIn(op["wallet"], active)
            self.assertGreater(Decimal(str(op["amount_btc"])), 0)
            if op["kind"] == "transfer":
                self.assertIn(op["to"], active)
                self.assertNotEqual(op["to"], op["wallet"])
            else:
                self.assertIsNone(op["to"])

    def test_tick_plan_requires_active_wallets(self):
        with self.assertRaises(ValueError):
            regtest_demo.plan_tick_operations([], random.Random(1))


if __name__ == "__main__":
    unittest.main()
